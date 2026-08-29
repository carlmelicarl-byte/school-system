#!/usr/bin/env python3
"""
ElimuPro School Management System — backend
A professional Zeraki / SmartShule-style school platform:
academics, analytics, finance, transport, attendance and communication.

Flask + SQLite, zero external services. Role-based access:
  admin    -> everything
  teacher  -> academics, marks, attendance, transport register
  accounts -> finance & fee management only (students read-only)
"""
import os
import re
import base64
import sqlite3
import datetime
import hashlib
import functools
import secrets
from flask import (Flask, g, jsonify, request, session, send_from_directory,
                   render_template)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "school.db")
UPLOAD_DIR = os.path.join(BASE_DIR, "static", "uploads")
DATA_DIR = os.path.join(BASE_DIR, "data")
META_DB = os.path.join(BASE_DIR, "meta.db")

# ------------------------------------------------------------------ multi-school
def open_meta():
    os.makedirs(DATA_DIR, exist_ok=True)
    m = sqlite3.connect(META_DB)
    m.row_factory = sqlite3.Row
    m.execute("""CREATE TABLE IF NOT EXISTS schools (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        slug TEXT UNIQUE NOT NULL,
        name TEXT NOT NULL,
        db_path TEXT NOT NULL,
        created_at TEXT DEFAULT (datetime('now')),
        active INTEGER DEFAULT 1)""")
    m.execute("""CREATE TABLE IF NOT EXISTS superusers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        full_name TEXT,
        active INTEGER DEFAULT 1)""")
    return m

def list_schools():
    m = open_meta()
    rows = m.execute("SELECT * FROM schools ORDER BY name").fetchall()
    m.close()
    return [dict(r) for r in rows]

def school_by_slug(slug):
    m = open_meta()
    r = m.execute("SELECT * FROM schools WHERE slug=?", (slug,)).fetchone()
    m.close()
    return dict(r) if r else None

def tenant_db_path(slug):
    s = school_by_slug(slug)
    if not s or not s.get("active"):
        return None
    p = s["db_path"]
    if not os.path.isabs(p):
        p = os.path.join(BASE_DIR, p)
    return p if os.path.exists(p) else None

def find_user_global(username):
    m = open_meta()
    su = m.execute("SELECT * FROM superusers WHERE username=? AND active=1", (username,)).fetchone()
    m.close()
    if su:
        return dict(su), "super"
    for sch in list_schools():
        if not sch.get("active"):
            continue
        try:
            c = sqlite3.connect(sch["db_path"])
            c.row_factory = sqlite3.Row
            u = c.execute("SELECT * FROM users WHERE username=? AND active=1", (username,)).fetchone()
            c.close()
        except Exception:
            continue
        if u:
            return dict(u), sch["slug"]
    return None, None

def ensure_platform():
    """On first run create the registry, super admin and a demo school so the
    platform is never empty. Existing data is never touched."""
    import hashlib as _hl
    m = open_meta()
    su = m.execute("SELECT id FROM superusers WHERE username='superadmin'").fetchone()
    if not su:
        salt = secrets.token_hex(16)
        h = _hl.scrypt(b"admin123", salt=salt.encode(), n=2 ** 14, r=8, p=1, dklen=32)
        m.execute("INSERT INTO superusers(username,password_hash,full_name) VALUES(?,?,?)",
                  ("superadmin", f"scrypt${salt}${h.hex()}", "Platform Administrator"))
        m.commit()
    demos = m.execute("SELECT slug, db_path FROM schools WHERE slug='greenfield'").fetchall()
    m.close()
    main_db = os.path.join(BASE_DIR, "school.db")
    if not os.path.exists(main_db):
        import seed as _seed
        print("[init] Creating demo school database...")
        _seed.seed_db(main_db, school_name="Greenfield Academy", sample=True,
                      admin_user="admin", admin_pass="admin123")
        _seed.register_school("greenfield", "Greenfield Academy", "school.db")
    elif not demos:
        import seed as _seed
        _seed.register_school("greenfield", "Greenfield Academy", "school.db")

def _host_slug():
    if getattr(g, "_host_slug", None) is not None:
        return g._host_slug
    host = (request.host or "").split(":")[0].lower()
    parts = host.split(".")
    slug = None
    if len(parts) >= 2 and parts[0] not in ("localhost", "www", "127"):
        cand = parts[0]
        if school_by_slug(cand):
            slug = cand
    g._host_slug = slug
    return slug

def _current_slug():
    hs = _host_slug()
    if hs:
        return hs
    if session.get("school_slug"):
        return session["school_slug"]
    hdr = request.headers.get("Authorization", "")
    if hdr.startswith("Bearer "):
        entry = AUTH_TOKENS.get(hdr[7:].strip())
        if entry and entry != "super":
            return entry[1]
    return None

def boot():
    ensure_platform()
app = Flask(__name__)

def _load_secret():
    """Secret key from ELIMUPRO_SECRET env var, else a generated file, else in-memory."""
    env = os.environ.get("ELIMUPRO_SECRET")
    if env:
        return env
    keyfile = os.path.join(BASE_DIR, ".secret_key")
    try:
        with open(keyfile) as f:
            k = f.read().strip()
        if k:
            return k
    except Exception:
        pass
    k = secrets.token_hex(32)
    try:
        with open(keyfile, "w") as f:
            f.write(k)
    except Exception:
        pass
    return k

app.secret_key = _load_secret()
app.config["JSON_SORT_KEYS"] = False
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
# set Secure cookie when deployed behind HTTPS (ELIMUPRO_HTTPS=1)
if os.environ.get("ELIMUPRO_HTTPS") == "1":
    app.config["SESSION_COOKIE_SECURE"] = True

# ---- login rate limiting (in-memory): 5 failures per user/IP per 15 min ----
from collections import defaultdict
import time as _time
FAIL_LIMIT = 5
FAIL_WINDOW = 900   # 15 minutes
_fails = defaultdict(list)

def _check_rate(uid):
    now = _time.time()
    _fails[uid] = [t for t in _fails[uid] if now - t < FAIL_WINDOW]
    if len(_fails[uid]) >= FAIL_LIMIT:
        wait = int(FAIL_WINDOW - (now - _fails[uid][0]))
        return f"Too many failed attempts. Try again in {max(1, wait // 60)} minute(s)."
    return None

def _record_fail(uid):
    _fails[uid].append(_time.time())

# In-memory bearer tokens: login returns a token that works even in sandboxed
# preview iframes where cookies are blocked. Token -> user id.
AUTH_TOKENS = {}

# ------------------------------------------------------------------ helpers
def db():
    if "db" not in g:
        slug = _current_slug()
        path = tenant_db_path(slug) if slug else DB_PATH
        g.db = sqlite3.connect(path)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db

@app.teardown_appcontext
def close_db(exc):
    d = g.pop("db", None)
    if d is not None:
        d.close()

def q(sql, args=()):
    return db().execute(sql, args).fetchall()

def q1(sql, args=()):
    return db().execute(sql, args).fetchone()

def exe(sql, args=()):
    cur = db().execute(sql, args)
    db().commit()
    return cur.lastrowid

def rows_to_dicts(rows):
    return [dict(r) for r in rows]

def phash(p, salt=None):
    """scrypt password hash, stored as scrypt$<salt>$<hash>."""
    if salt is None:
        salt = secrets.token_hex(16)
    h = hashlib.scrypt(p.encode(), salt=salt.encode(), n=2 ** 14, r=8, p=1, dklen=32)
    return f"scrypt${salt}${h.hex()}"

def verify_password(stored, p):
    """Verify a password against a stored hash (scrypt, with legacy SHA-256 support)."""
    if not stored:
        return False
    if stored.startswith("scrypt$"):
        try:
            _, salt, hexhash = stored.split("$")
            h = hashlib.scrypt(p.encode(), salt=salt.encode(), n=2 ** 14, r=8, p=1, dklen=32)
            return secrets.compare_digest(h.hex(), hexhash)
        except Exception:
            return False
    # legacy SHA-256 hash from older versions
    return secrets.compare_digest(hashlib.sha256(p.encode()).hexdigest(), stored)

def _num_words(n):
    ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine", "Ten",
            "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
    if n == 0:
        return "Zero"
    def two(x):
        return ones[x] if x < 20 else tens[x // 10] + ("-" + ones[x % 10] if x % 10 else "")
    def three(x):
        return (ones[x // 100] + " Hundred" + (" and " + two(x % 100) if x % 100 else "") if x >= 100 else two(x))
    words, i = "", 0
    scale = ["", " Thousand", " Million", " Billion"]
    while n > 0:
        part = n % 1000
        if part:
            words = three(part) + scale[i] + (" " + words if words else "")
        n //= 1000
        i += 1
    return words

def amount_in_words(amount):
    sh = int(float(amount or 0))
    cents = round((float(amount or 0) - sh) * 100)
    w = "Kenya Shillings " + _num_words(sh)
    if cents > 0:
        w += " and " + _num_words(cents) + " Cents"
    return w + " Only"

def fmt_amount(n):
    s = settings_map()
    cur = s.get("currency", "KSh")
    return f"{cur} {float(n or 0):,.0f}"

def auth_user():
    """Current user row from cookie session or Bearer token (superadmin has no tenant)."""
    def super_row(uid, uname, full):
        return {"id": uid, "username": uname, "full_name": full, "role": "superadmin"}
    if session.get("user_id"):
        if session.get("is_super"):
            return super_row(session["user_id"], session.get("username"), session.get("name"))
        return q1("SELECT * FROM users WHERE id=?", (session["user_id"],))
    hdr = request.headers.get("Authorization", "")
    if hdr.startswith("Bearer "):
        entry = AUTH_TOKENS.get(hdr[7:].strip())
        if entry == "super":
            return super_row(0, "superadmin", "Platform Administrator")
        if entry:
            return q1("SELECT * FROM users WHERE id=?", (entry[0],))
    return None

def acting_name():
    u = auth_user()
    return u["full_name"] if u else "Admin"

def link_students(uid, student_ids):
    """Replace the set of students linked to a guardian account."""
    cur = db()
    cur.execute("DELETE FROM guardian_links WHERE user_id=?", (uid,))
    seen = set()
    for sid in student_ids:
        sid = int(sid)
        if sid in seen:
            continue
        seen.add(sid)
        if q1("SELECT id FROM students WHERE id=?", (sid,)):
            cur.execute("INSERT INTO guardian_links(user_id,student_id) VALUES(?,?)", (uid, sid))
    cur.commit()

def log_activity(action, detail=""):
    """Record an audit entry for the current user."""
    u = auth_user()
    try:
        exe("INSERT INTO activity_log(user_id,action,detail) VALUES(?,?,?)",
            (u["id"] if u else None, action, (detail or "")[:250]))
    except Exception:
        pass

def role_required(*roles):
    def deco(fn):
        @functools.wraps(fn)
        def wrap(*a, **kw):
            u = auth_user()
            if not u:
                return jsonify({"error": "Unauthorized"}), 401
            if u["role"] not in roles:
                return jsonify({"error": "Forbidden"}), 403
            return fn(*a, **kw)
        return wrap
    return deco

login_required = role_required("admin", "teacher", "accounts", "librarian")
any_required = role_required("admin", "teacher", "accounts", "guardian", "librarian")
admin_required = role_required("admin")
finance_required = role_required("admin", "accounts")
academic_required = role_required("admin", "teacher")
guardian_required = role_required("guardian")
library_required = role_required("admin", "librarian")

def super_required(fn):
    @functools.wraps(fn)
    def wrap(*a, **kw):
        u = auth_user()
        if not u:
            return jsonify({"error": "Unauthorized"}), 401
        if u["role"] != "superadmin":
            return jsonify({"error": "Platform admin only"}), 403
        return fn(*a, **kw)
    return wrap


# ------------------------------------------------------------------ grading
# Kenyan CBC (Competency-Based Curriculum) achievement levels — used for ALL
# grades (primary, junior secondary and senior secondary):
#   E = Exceeding Expectations (80%+), M = Meeting (65%+),
#   A = Approaching (50%+), B = Below Expectations (<50%)
CBC_BANDS = [
    (80, "E", "Exceeding Expectations", 4),
    (65, "M", "Meeting Expectations", 3),
    (50, "A", "Approaching Expectations", 2),
    (0,  "B", "Below Expectations", 1),
]

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
PERIODS = [
    {"n": 1, "start": "8:00", "end": "8:40"},
    {"n": 2, "start": "8:40", "end": "9:20"},
    {"n": 3, "start": "9:20", "end": "10:00"},
    {"n": 4, "start": "10:20", "end": "11:00"},
    {"n": 5, "start": "11:00", "end": "11:40"},
    {"n": 6, "start": "11:40", "end": "12:20"},
    {"n": 7, "start": "13:20", "end": "14:00"},
    {"n": 8, "start": "14:00", "end": "14:40"},
]

def scale_for_grade(grade_str):
    """The whole school uses CBC achievement levels (E/M/A/B) in every grade."""
    return "cbc"

def grade_for(score, scale="cbc"):
    if score is None:
        return None, None
    for lo, letter, _name, pts in CBC_BANDS:
        if score >= lo:
            return letter, pts
    return "B", 1

def level_name(letter, scale="cbc"):
    for _lo, l, name, _pts in CBC_BANDS:
        if l == letter:
            return name
    return letter

def mean_grade_from_points(avg_pts, scale="cbc"):
    if avg_pts is None:
        return "-"
    if avg_pts >= 3.5: return "E"
    if avg_pts >= 2.5: return "M"
    if avg_pts >= 1.5: return "A"
    return "B"

def conduct_rating(net):
    """CBC-style conduct rating from net merits (merits - demerits)."""
    if net >= 6: return "Excellent"
    if net >= 2: return "Good"
    if net >= -1: return "Satisfactory"
    return "Needs Improvement"

def subjects_for_grade(gnum):
    """Subjects taught in a given grade (CBC curriculum)."""
    rows = q("SELECT * FROM subjects")
    out = []
    for r in rows:
        try:
            gs = [int(x) for x in (r["grades"] or "").split(",") if x.strip().isdigit()]
        except Exception:
            gs = []
        if gnum in gs or not r["grades"]:
            out.append(r)
    return out

def settings_map():
    return dict((r["key"], r["value"]) for r in q("SELECT key,value FROM settings"))

def student_class(student_id, term=None):
    s = settings_map()
    term = term or s.get("current_term", "Term 3")
    r = q1("""SELECT c.* FROM enrollments e JOIN classes c ON c.id=e.class_id
              WHERE e.student_id=? AND e.term=? AND e.academic_year=? ORDER BY e.id DESC LIMIT 1""",
           (student_id, term, s.get("academic_year", "2026")))
    return dict(r) if r else None

def students_with_class(term=None):
    s = settings_map()
    term = term or s.get("current_term", "Term 3")
    year = s.get("academic_year", "2026")
    rows = q("""SELECT st.*, c.id AS class_id, c.name AS class_name, c.grade AS grade
                FROM students st
                LEFT JOIN enrollments e ON e.student_id=st.id AND e.term=? AND e.academic_year=?
                LEFT JOIN classes c ON c.id=e.class_id
                ORDER BY st.first_name""", (term, year))
    return rows_to_dicts(rows)

def exam_results(exam_id):
    rows = q("""SELECT student_id, COUNT(es.id) AS subjects,
                       SUM(es.points) AS total_points,
                       ROUND(AVG(es.score),1) AS mean,
                       ROUND(AVG(es.points),2) AS avg_pts
                FROM exam_scores es WHERE es.exam_id=?
                GROUP BY es.student_id""", (exam_id,))
    return {r["student_id"]: dict(r) for r in rows}

# ---------------------------------------------------------------- billing
def billing_rows(term=None, year=None, student_id=None, class_id=None):
    """Tuition + transport billed per student. Returns [(student_id, billed)]."""
    s = settings_map()
    term = term or s.get("current_term", "Term 3")
    year = year or s.get("academic_year", "2026")
    sql = """SELECT e.student_id,
                    fs.amount + COALESCE(tr.fee, 0) AS billed
             FROM enrollments e
             JOIN fee_structures fs
               ON fs.class_id=e.class_id AND fs.term=e.term AND fs.academic_year=e.academic_year
             LEFT JOIN transport_assignments ta
               ON ta.student_id=e.student_id AND ta.academic_year=e.academic_year AND ta.status='Active'
             LEFT JOIN transport_routes tr ON tr.id=ta.route_id
             WHERE e.term=? AND e.academic_year=?"""
    args = [term, year]
    if student_id:
        sql += " AND e.student_id=?"
        args.append(student_id)
    if class_id:
        sql += " AND e.class_id=?"
        args.append(class_id)
    rows = q(sql, args)
    return [(r["student_id"], r["billed"]) for r in rows]

def billed_for(student_id, term=None, year=None):
    rows = billing_rows(term=term, year=year, student_id=student_id)
    return rows[0][1] if rows else 0

def paid_for(student_id, term, year):
    return q1("""SELECT COALESCE(SUM(fp.amount),0) a FROM fee_payments fp
                 WHERE fp.student_id=? AND fp.term=? AND fp.payment_date LIKE ?""",
              (student_id, term, f"{year}-%"))["a"]

def finance_snapshot(term, year):
    classes = q("SELECT c.id, c.name FROM classes c WHERE c.academic_year=? ORDER BY c.name", (year,))
    class_rows, total_billed, total_paid = [], 0, 0
    for c in classes:
        billed = sum(b for sid, b in billing_rows(term=term, year=year, class_id=c["id"]))
        paid = q1("""SELECT COALESCE(SUM(fp.amount),0) a FROM fee_payments fp
                     WHERE fp.term=? AND fp.payment_date LIKE ?
                     AND fp.student_id IN (SELECT student_id FROM enrollments WHERE class_id=?)""",
                  (term, f"{year}-%", c["id"]))["a"]
        total_billed += billed
        total_paid += paid
        class_rows.append({"id": c["id"], "name": c["name"], "billed": billed, "paid": paid,
                           "arrears": billed - paid,
                           "rate": round(paid / billed * 100, 1) if billed else 0})
    return total_billed, total_paid, class_rows

# ------------------------------------------------------------------ pages
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/static/<path:path>")
def static_files(path):
    return send_from_directory(os.path.join(BASE_DIR, "static"), path)

@app.route("/sw.js")
def service_worker():
    resp = send_from_directory(os.path.join(BASE_DIR, "static"), "sw.js")
    resp.headers["Service-Worker-Allowed"] = "/"
    resp.headers["Cache-Control"] = "no-cache"
    return resp

@app.after_request
def security_headers(resp):
    resp.headers.setdefault("X-Content-Type-Options", "nosniff")
    resp.headers.setdefault("Referrer-Policy", "same-origin")
    resp.headers.setdefault("X-XSS-Protection", "0")
    if os.environ.get("ELIMUPRO_HTTPS") == "1":
        resp.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")
    return resp

@app.errorhandler(404)
def not_found(e):
    if request.path.startswith("/api/"):
        return jsonify({"error": "Not found"}), 404
    return render_template("index.html"), 404

# ------------------------------------------------------------------ auth
@app.route("/api/login", methods=["POST"])
def login():
    data = request.get_json(force=True) or {}
    uname = data.get("username", "").strip()
    ip = request.remote_addr or "?"
    key = f"{uname}|{ip}"
    blocked = _check_rate(key)
    if blocked:
        return jsonify({"error": blocked}), 429
    host_slug = _host_slug()
    if host_slug:
        sch = school_by_slug(host_slug)
        u = None
        if sch and sch.get("active"):
            try:
                c = sqlite3.connect(sch["db_path"])
                c.row_factory = sqlite3.Row
                u = c.execute("SELECT * FROM users WHERE username=? AND active=1", (uname,)).fetchone()
                c.close()
            except Exception:
                u = None
        slug = host_slug if u else None
    else:
        u, slug = find_user_global(uname)
    if not u or not verify_password(u["password_hash"], data.get("password", "")):
        _record_fail(key)
        return jsonify({"error": "Wrong password or username. Please check your details and try again."}), 401
    _fails.pop(key, None)
    token = secrets.token_hex(16)
    if slug == "super":
        session["user_id"] = 0
        session["is_super"] = True
        session["username"] = u["username"]
        session["name"] = u["full_name"]
        session.pop("school_slug", None)
        AUTH_TOKENS[token] = "super"
        return jsonify({"id": 0, "username": u["username"], "full_name": u["full_name"],
                        "role": "superadmin", "token": token})
    # school user: upgrade legacy hashes + bind token/session to the school
    if u["password_hash"] and not u["password_hash"].startswith("scrypt$"):
        c = sqlite3.connect(tenant_db_path(slug))
        c.execute("UPDATE users SET password_hash=? WHERE id=?", (phash(data.get("password", "")), u["id"]))
        c.commit(); c.close()
    session["user_id"] = u["id"]
    session["role"] = u["role"]
    session["name"] = u["full_name"]
    session["school_slug"] = slug
    AUTH_TOKENS[token] = (u["id"], slug)
    return jsonify({"id": u["id"], "username": u["username"], "full_name": u["full_name"],
                    "role": u["role"], "teacher_id": u["teacher_id"],
                    "profile_pic": u["profile_pic"], "token": token, "school_slug": slug})

@app.route("/api/logout", methods=["POST"])
def logout():
    hdr = request.headers.get("Authorization", "")
    if hdr.startswith("Bearer "):
        AUTH_TOKENS.pop(hdr[7:].strip(), None)
    session.clear()
    return jsonify({"ok": True})

# ------------------------------------------------------------------ forgot password
def _ensure_reset_table():
    db().execute("""CREATE TABLE IF NOT EXISTS reset_tokens (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        code_hash TEXT NOT NULL,
        expires_at TEXT NOT NULL,
        used INTEGER DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now')))""")
    db().commit()

@app.route("/api/auth/forgot", methods=["POST"])
def forgot_password():
    """Request a password-reset code. 'Sends' a 6-digit code via the SMS log
    (simulated — in production this goes through the SMS gateway) and stores a
    hashed copy with a 15-minute expiry. Returns demo_code only because SMS is
    simulated here; a real gateway would never return it."""
    d = request.get_json(force=True) or {}
    uname = d.get("username", "").strip()
    _ensure_reset_table()
    code = None
    if uname:
        u = q1("SELECT * FROM users WHERE username=?", (uname,))
        if u:
            code = f"{secrets.randbelow(1000000):06d}"
            expires = (datetime.datetime.now() + datetime.timedelta(minutes=15)).isoformat()
            exe("INSERT INTO reset_tokens(user_id,code_hash,expires_at) VALUES(?,?,?)",
                (u["id"], phash(code), expires))
            # resolve a phone number to 'send' to (guardian username IS the phone)
            phone = ""
            if u["role"] == "guardian" and uname.isdigit():
                phone = uname
            elif u["role"] == "teacher" and u["teacher_id"]:
                tr = q1("SELECT phone FROM teachers WHERE id=?", (u["teacher_id"],))
                phone = tr["phone"] if tr and tr["phone"] else phone
            elif u["role"] == "librarian":
                tr = q1("SELECT phone FROM teachers WHERE id=?", (u["teacher_id"],))
                phone = tr["phone"] if tr and tr["phone"] else phone
            exe("""INSERT INTO sms_log(to_phone,parent_name,student_name,message,category,status)
                   VALUES(?,?,?,?,?,?)""",
                (phone, u["full_name"], "",
                 f"Your ElimuPro password reset code is {code}. It expires in 15 minutes. - {settings_map().get('school_name','School')}",
                 "Password Reset", "Sent"))
    # never reveal whether the username exists
    return jsonify({"ok": True, "demo_code": code if code else None})

@app.route("/api/auth/reset", methods=["POST"])
def reset_password():
    """Verify the code and set a new password (one step)."""
    d = request.get_json(force=True) or {}
    uname = d.get("username", "").strip()
    code = str(d.get("code", "")).strip()
    newpass = d.get("new_password", "")
    if len(newpass) < 6:
        return jsonify({"error": "Password must be at least 6 characters"}), 400
    u = q1("SELECT * FROM users WHERE username=?", (uname,))
    if not u:
        return jsonify({"error": "Invalid reset details"}), 400
    _ensure_reset_table()
    now = datetime.datetime.now().isoformat()
    tok = q1("""SELECT * FROM reset_tokens WHERE user_id=? AND used=0 AND expires_at>? ORDER BY id DESC LIMIT 1""",
             (u["id"], now))
    if not tok or not verify_password(tok["code_hash"], code):
        return jsonify({"error": "Invalid or expired code. Please request a new one."}), 400
    cur = db()
    cur.execute("UPDATE reset_tokens SET used=1 WHERE id=?", (tok["id"],))
    cur.execute("UPDATE users SET password_hash=? WHERE id=?", (phash(newpass), u["id"]))
    cur.commit()
    log_activity("Password reset", f"{uname} reset their password")
    return jsonify({"ok": True})

@app.route("/api/me")
@any_required
def me():
    u = auth_user()
    return jsonify({"id": u["id"], "name": u["full_name"], "role": u["role"],
                    "username": u["username"], "profile_pic": u["profile_pic"],
                    "teacher_id": (u["teacher_id"] if "teacher_id" in u else None),
                    "school_slug": _current_slug()})

# ------------------------------------------------------------------ users
@app.route("/api/users")
@admin_required
def list_users():
    rows = q("""SELECT u.id, u.username, u.full_name, u.role, u.active, u.profile_pic,
                       t.first_name, t.last_name,
                       (SELECT COUNT(*) FROM guardian_links gl WHERE gl.user_id=u.id) child_count
                FROM users u LEFT JOIN teachers t ON t.id=u.teacher_id
                ORDER BY u.role, u.username""")
    return jsonify(rows_to_dicts(rows))

@app.route("/api/users", methods=["POST"])
@admin_required
def add_user():
    d = request.get_json(force=True) or {}
    if not d.get("username") or not d.get("password") or not d.get("full_name"):
        return jsonify({"error": "username, password and full_name required"}), 400
    if d.get("role") not in ("admin", "teacher", "accounts", "guardian"):
        return jsonify({"error": "Invalid role"}), 400
    try:
        uid = exe("INSERT INTO users(username,password_hash,full_name,role) VALUES(?,?,?,?)",
                  (d["username"].strip(), phash(d["password"]), d["full_name"].strip(), d["role"]))
    except sqlite3.IntegrityError:
        return jsonify({"error": "Username already exists"}), 400
    # guardian accounts can be linked to one or more students (shared portal)
    if d.get("role") == "guardian" and d.get("student_ids"):
        link_students(uid, d["student_ids"])
    log_activity("User created", f"{d['username']} ({d['role']})")
    return jsonify({"id": uid})

@app.route("/api/users/<int:uid>", methods=["PUT"])
@admin_required
def update_user(uid):
    d = request.get_json(force=True) or {}
    fields = ["full_name", "role", "active"]
    sets = [f for f in fields if f in d]
    if "role" in d and d["role"] not in ("admin", "teacher", "accounts", "guardian"):
        return jsonify({"error": "Invalid role"}), 400
    if sets:
        exe("UPDATE users SET " + ", ".join(f"{f}=?" for f in sets) + " WHERE id=?",
            tuple(d[f] for f in sets) + (uid,))
    # (re)link children for guardian accounts
    if "student_ids" in d and d.get("student_ids") is not None:
        link_students(uid, d["student_ids"])
    log_activity("User updated", f"#{uid}")
    return jsonify({"ok": True})

@app.route("/api/users/<int:uid>/children")
@admin_required
def user_children(uid):
    rows = q("""SELECT st.id, st.first_name, st.last_name, st.admission_no, st.profile_pic,
                       c.name class_name
                FROM guardian_links gl
                JOIN students st ON st.id=gl.student_id
                LEFT JOIN enrollments e ON e.student_id=st.id AND e.term=(SELECT value FROM settings WHERE key='current_term')
                LEFT JOIN classes c ON c.id=e.class_id
                WHERE gl.user_id=? ORDER BY st.first_name""", (uid,))
    return jsonify(rows_to_dicts(rows))

@app.route("/api/users/<int:uid>/password", methods=["PUT"])
@admin_required
def reset_user_password(uid):
    d = request.get_json(force=True) or {}
    if not d.get("password"):
        return jsonify({"error": "password required"}), 400
    exe("UPDATE users SET password_hash=? WHERE id=?", (phash(d["password"]), uid))
    return jsonify({"ok": True})

@app.route("/api/me/password", methods=["PUT"])
@any_required
def change_my_password():
    d = request.get_json(force=True) or {}
    u = auth_user()
    if u["password_hash"] != phash(d.get("old", "")):
        return jsonify({"error": "Current password is incorrect"}), 400
    if not d.get("new"):
        return jsonify({"error": "New password required"}), 400
    exe("UPDATE users SET password_hash=? WHERE id=?", (phash(d["new"]), u["id"]))
    return jsonify({"ok": True})

# ------------------------------------------------------------------ uploads
@app.route("/api/upload/<kind>/<int:rid>", methods=["POST"])
@any_required
def upload_pic(kind, rid):
    u = auth_user()
    if kind == "user":
        if u["id"] != rid and u["role"] != "admin":
            return jsonify({"error": "Forbidden"}), 403
    elif kind in ("student", "teacher", "school"):
        if u["role"] != "admin":
            return jsonify({"error": "Forbidden"}), 403
    else:
        return jsonify({"error": "Invalid kind"}), 400

    data = (request.get_json(force=True) or {}).get("data", "")
    m = re.match(r"data:image/(png|jpe?g|gif|webp);base64,(.+)", data)
    if not m:
        return jsonify({"error": "Invalid image data"}), 400
    ext = {"png": "png", "jpg": "jpg", "jpeg": "jpg", "gif": "gif", "webp": "webp"}[m.group(1)]
    try:
        raw = base64.b64decode(m.group(2))
    except Exception:
        return jsonify({"error": "Invalid image data"}), 400
    if len(raw) > 4 * 1024 * 1024:
        return jsonify({"error": "Image too large (max 4MB)"}), 400
    subdir = _current_slug() or "default"
    target = os.path.join(UPLOAD_DIR, subdir)
    os.makedirs(target, exist_ok=True)
    fname = f"{kind}_{rid}_{secrets.token_hex(4)}.{ext}"
    with open(os.path.join(target, fname), "wb") as f:
        f.write(raw)
    path = f"/static/uploads/{subdir}/{fname}"
    if kind == "user":
        exe("UPDATE users SET profile_pic=? WHERE id=?", (path, rid))
    elif kind == "student":
        exe("UPDATE students SET profile_pic=? WHERE id=?", (path, rid))
    elif kind == "teacher":
        exe("UPDATE teachers SET profile_pic=? WHERE id=?", (path, rid))
    else:
        exe("INSERT INTO settings(key,value) VALUES('school_logo',?) ON CONFLICT(key) DO UPDATE SET value=excluded.value", (path,))
    return jsonify({"path": path})

# ------------------------------------------------------------------ dashboard
@app.route("/api/dashboard")
@login_required
def dashboard():
    s = settings_map()
    term = s.get("current_term", "Term 3")
    year = s.get("academic_year", "2026")

    total_students = q1("SELECT COUNT(*) c FROM students WHERE status='Active'")["c"]
    total_teachers = q1("SELECT COUNT(*) c FROM teachers WHERE active=1")["c"]
    total_classes = q1("SELECT COUNT(*) c FROM classes WHERE academic_year=?", (year,))["c"]
    total_subjects = q1("SELECT COUNT(*) c FROM subjects")["c"]

    billed = sum(b for _, b in billing_rows(year=year)) * 3  # 3 terms per year
    billed = sum(b for _, b in [x for t in ("Term 1", "Term 2", "Term 3") for x in billing_rows(term=t, year=year)])
    paid = q1("""SELECT COALESCE(SUM(fp.amount),0) a FROM fee_payments fp
                 JOIN students st ON st.id=fp.student_id WHERE st.status='Active'
                 AND fp.payment_date LIKE ?""", (f"{year}-%",))["a"]

    paid_term = q1("""SELECT COALESCE(SUM(amount),0) a FROM fee_payments
                      WHERE term=? AND payment_date LIKE ?""", (term, f"{year}-%"))["a"]
    billed_term = sum(b for _, b in billing_rows(term=term, year=year))

    fee_breakdown = []
    for c in q("SELECT c.id, c.name FROM classes c WHERE c.academic_year=? ORDER BY c.name", (year,)):
        b = sum(x for _, x in billing_rows(year=year, class_id=c["id"])) * 3
        p = q1("""SELECT COALESCE(SUM(fp.amount),0) a FROM fee_payments fp
                  JOIN students st ON st.id=fp.student_id
                  JOIN enrollments e ON e.student_id=st.id AND e.academic_year=?
                  WHERE e.class_id=? AND fp.payment_date LIKE ?""", (year, c["id"], f"{year}-%"))["a"]
        fee_breakdown.append({"id": c["id"], "name": c["name"], "billed": b, "paid": p})

    att_today = q1("SELECT COUNT(*) c FROM attendance WHERE date=?", (datetime.date.today().isoformat(),))["c"]

    closed = q1("SELECT * FROM exams WHERE status='Closed' ORDER BY id DESC LIMIT 1")
    perf = None
    if closed:
        res = q("SELECT ROUND(AVG(score),1) mean FROM exam_scores WHERE exam_id=?", (closed["id"],))
        grade_dist = q("SELECT grade, COUNT(*) c FROM exam_scores WHERE exam_id=? GROUP BY grade", (closed["id"],))
        perf = {
            "exam_id": closed["id"], "exam_name": closed["name"], "term": closed["term"],
            "mean": res[0]["mean"] if res else 0,
            "grade_dist": {r["grade"]: r["c"] for r in grade_dist},
            "subject_means": rows_to_dicts(q("""SELECT su.id, su.name, ROUND(AVG(es.score),1) mean,
                                                       MAX(es.score) highest, MIN(es.score) lowest
                                                FROM exam_scores es JOIN subjects su ON su.id=es.subject_id
                                                WHERE es.exam_id=? GROUP BY su.id ORDER BY mean DESC""", (closed["id"],))),
            "class_means": rows_to_dicts(q("""SELECT c.id, c.name, ROUND(AVG(es.score),1) mean, COUNT(DISTINCT es.student_id) students
                                               FROM exam_scores es
                                               JOIN enrollments e ON e.student_id=es.student_id AND e.term=? AND e.academic_year=?
                                               JOIN classes c ON c.id=e.class_id
                                               WHERE es.exam_id=? GROUP BY c.id ORDER BY mean DESC""", (term, year, closed["id"]))),
        }

    recent_payments = rows_to_dicts(q("""SELECT fp.*, st.first_name, st.last_name, st.admission_no
                                         FROM fee_payments fp JOIN students st ON st.id=fp.student_id
                                         ORDER BY fp.payment_date DESC, fp.id DESC LIMIT 6"""))
    recent_ann = rows_to_dicts(q("SELECT * FROM announcements ORDER BY id DESC LIMIT 4"))
    active_exam = q1("SELECT * FROM exams WHERE status='Open' ORDER BY id DESC LIMIT 1")

    # transport summary for dashboard
    tr_summary = {
        "routes": q1("SELECT COUNT(*) c FROM transport_routes WHERE status='Active'")["c"],
        "assigned": q1("SELECT COUNT(*) c FROM transport_assignments WHERE status='Active' AND academic_year=?", (year,))["c"],
        "boarded_today": q1("SELECT COUNT(*) c FROM transport_log WHERE date=? AND status='Boarded'",
                            (datetime.date.today().isoformat(),))["c"],
        "monthly_fees": q1("""SELECT COALESCE(SUM(tr.fee),0) a FROM transport_assignments ta
                              JOIN transport_routes tr ON tr.id=ta.route_id
                              WHERE ta.status='Active' AND ta.academic_year=?""", (year,))["a"],
    }

    # attention list — things that need action
    unplaced = q1("""SELECT COUNT(*) c FROM students st
                     WHERE st.status='Active' AND NOT EXISTS (
                       SELECT 1 FROM enrollments e
                       WHERE e.student_id=st.id AND e.term=? AND e.academic_year=?)""",
                  (term, year))["c"]
    open_exams = q1("SELECT COUNT(*) c FROM exams WHERE status='Open'")["c"]
    critical = 0
    for st in students_with_class(term):
        b = billed_for(st["id"], term, year)
        p = paid_for(st["id"], term, year)
        if b and (b - p) >= b * 0.5:
            critical += 1
    role = auth_user()["role"]
    alerts = []
    if role != "accounts":
        alerts.append({"icon": "i-users", "label": "Unplaced students", "value": unplaced, "view": "students",
                       "tone": "amber" if unplaced else "green"})
        alerts.append({"icon": "i-exam", "label": "Open exam in progress", "value": open_exams, "view": "exams",
                       "tone": "blue" if open_exams else "green"})
    alerts.append({"icon": "i-money", "label": "Critical fee balances", "value": critical, "view": "finance",
                   "tone": "red" if critical else "green"})
    feed = rows_to_dicts(q("""SELECT a.*, u.full_name user_name, u.role user_role
                              FROM activity_log a LEFT JOIN users u ON u.id=a.user_id
                              ORDER BY a.id DESC LIMIT 8"""))

    lib = {
        "total_titles": q1("SELECT COUNT(*) c FROM books")["c"],
        "issued": q1("SELECT COUNT(*) c FROM book_issues WHERE status IN ('Issued','Overdue')")["c"],
        "overdue": q1("SELECT COUNT(*) c FROM book_issues WHERE status='Overdue'")["c"],
    }
    tdy = datetime.date.today().isoformat()
    upcoming_events = rows_to_dicts(q("""SELECT * FROM school_events WHERE event_date>=? ORDER BY event_date LIMIT 4""", (tdy,)))
    conduct = {"merits": q1("SELECT COUNT(*) c FROM conduct_records WHERE record_type='Merit' AND record_date>=?", (s.get("term_start") or "2026-05-01",))["c"],
               "demerits": q1("SELECT COUNT(*) c FROM conduct_records WHERE record_type='Demerit' AND record_date>=?", (s.get("term_start") or "2026-05-01",))["c"]}
    due_hw = rows_to_dicts(q("""SELECT hw.*, c.name class_name, s.name subject_name
                                FROM homework hw
                                JOIN classes c ON c.id=hw.class_id
                                LEFT JOIN subjects s ON s.id=hw.subject_id
                                WHERE hw.due_date>=? ORDER BY hw.due_date LIMIT 5""", (tdy,)))
    return jsonify({
        "settings": s,
        "library": lib,
        "upcoming_events": upcoming_events,
        "conduct": conduct,
        "due_homework": due_hw,
        "alerts": alerts,
        "activity": feed,
        "counts": {"students": total_students, "teachers": total_teachers,
                   "classes": total_classes, "subjects": total_subjects},
        "finance": {"billed": billed, "paid": paid, "arrears": billed - paid,
                    "billed_term": billed_term, "paid_term": paid_term,
                    "arrears_term": billed_term - paid_term, "class_breakdown": fee_breakdown},
        "attendance_today": att_today,
        "performance": perf,
        "transport": tr_summary,
        "active_exam": dict(active_exam) if active_exam else None,
        "recent_payments": recent_payments,
        "recent_announcements": recent_ann,
        "term": term, "year": year,
    })

# ------------------------------------------------------------------ students
@app.route("/api/students")
@login_required
def list_students():
    q_ = request.args.get("q", "").strip().lower()
    class_id = request.args.get("class_id")
    status = request.args.get("status", "")
    rows = students_with_class()
    if q_:
        rows = [r for r in rows if q_ in (r["first_name"] + " " + r.get("middle_name", "") + " " + r["last_name"]).lower()
                or q_ in r["admission_no"].lower() or q_ in (r.get("parent_name") or "").lower()]
    if class_id:
        rows = [r for r in rows if str(r.get("class_id")) == class_id]
    if status:
        rows = [r for r in rows if r["status"] == status]
    return jsonify(rows)

@app.route("/api/students", methods=["POST"])
@admin_required
def add_student():
    d = request.get_json(force=True) or {}
    need = ["first_name", "last_name", "gender"]
    if any(not d.get(k) for k in need):
        return jsonify({"error": "first_name, last_name and gender are required"}), 400
    year = settings_map().get("academic_year", "2026")
    adm = (d.get("admission_no") or "").strip() or f"GF/{year}/{datetime.date.today().year%100}{random_seq()}"
    try:
        st_id = exe("""INSERT INTO students(admission_no,first_name,middle_name,last_name,gender,dob,admission_date,
                                            parent_name,parent_phone,parent_email,address,status,blood_group,house)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (adm, d.get("first_name"), d.get("middle_name", ""), d.get("last_name"),
                     d.get("gender"), d.get("dob"), d.get("admission_date") or datetime.date.today().isoformat(),
                     d.get("parent_name"), d.get("parent_phone"), d.get("parent_email"),
                     d.get("address"), "Active", d.get("blood_group"), d.get("house")))
    except sqlite3.IntegrityError:
        return jsonify({"error": "Admission number already exists"}), 400
    if d.get("class_id"):
        exe("INSERT INTO enrollments(student_id,class_id,term,academic_year) VALUES(?,?,?,?)",
            (st_id, d["class_id"], settings_map().get("current_term", "Term 3"), year))
    log_activity("Student admitted", f"{d['first_name']} {d['last_name']} ({adm})")
    return jsonify({"id": st_id})

def random_seq():
    import random
    return random.randint(1000, 9999)

@app.route("/api/students/<int:sid>", methods=["PUT"])
@admin_required
def update_student(sid):
    d = request.get_json(force=True) or {}
    fields = ["first_name", "middle_name", "last_name", "gender", "dob", "admission_date",
              "parent_name", "parent_phone", "parent_email", "address", "status", "admission_no",
              "blood_group", "house"]
    sets = [f for f in fields if f in d]
    if not sets:
        return jsonify({"error": "No fields"}), 400
    sql = "UPDATE students SET " + ", ".join(f"{f}=?" for f in sets) + " WHERE id=?"
    exe(sql, tuple(d[f] for f in sets) + (sid,))
    if d.get("class_id"):
        term = settings_map().get("current_term", "Term 3")
        year = settings_map().get("academic_year", "2026")
        cur = db()
        cur.execute("DELETE FROM enrollments WHERE student_id=? AND term=?", (sid, term))
        cur.execute("INSERT INTO enrollments(student_id,class_id,term,academic_year) VALUES(?,?,?,?)",
                    (sid, d["class_id"], term, year))
        cur.commit()
    log_activity("Student updated", f"#{sid} record edited")
    return jsonify({"ok": True})

@app.route("/api/students/<int:sid>")
@login_required
def student_detail(sid):
    st = q1("SELECT * FROM students WHERE id=?", (sid,))
    if not st:
        return jsonify({"error": "Not found"}), 404
    out = dict(st)
    out["class"] = student_class(sid)
    out["history"] = rows_to_dicts(q("""SELECT e.term, e.academic_year, c.name class_name
                                        FROM enrollments e JOIN classes c ON c.id=e.class_id
                                        WHERE e.student_id=? ORDER BY e.academic_year DESC, e.term""", (sid,)))
    out["payments"] = rows_to_dicts(q("""SELECT fp.*, pt.name payment_type_name, pt.category payment_type_category
                                         FROM fee_payments fp
                                         LEFT JOIN payment_types pt ON pt.id=fp.payment_type_id
                                         WHERE fp.student_id=? ORDER BY fp.payment_date DESC""", (sid,)))
    out["transport"] = rows_to_dicts(q("""SELECT tr.* FROM transport_assignments ta
                                          JOIN transport_routes tr ON tr.id=ta.route_id
                                          WHERE ta.student_id=? AND ta.status='Active'""", (sid,)))
    billing = {}
    for term in ("Term 1", "Term 2", "Term 3"):
        b = billed_for(sid, term, settings_map().get("academic_year", "2026"))
        if b:
            billing[term] = b
    out["billing"] = billing
    return jsonify(out)

@app.route("/api/students/promote", methods=["POST"])
@admin_required
def promote_class():
    """Move EVERY student of one class to another class for the WHOLE academic year.

    The source class is left completely empty (all three terms' enrollments are
    reassigned) so it is ready for a fresh intake — no mix-ups between classes.
    """
    d = request.get_json(force=True) or {}
    from_id, to_id = d.get("from_class_id"), d.get("to_class_id")
    if not from_id or not to_id:
        return jsonify({"error": "from_class_id and to_class_id required"}), 400
    if from_id == to_id:
        return jsonify({"error": "Source and target class are the same"}), 400
    s = settings_map()
    year = s.get("academic_year", "2026")
    from_c = q1("SELECT * FROM classes WHERE id=?", (from_id,))
    to_c = q1("SELECT * FROM classes WHERE id=?", (to_id,))
    if not from_c or not to_c:
        return jsonify({"error": "Class not found"}), 404
    rows = q("""SELECT DISTINCT student_id FROM enrollments
                WHERE class_id=? AND academic_year=?""", (from_id, year))
    if not rows:
        return jsonify({"error": "No students in the source class"}), 400
    cur = db()
    moved = 0
    for r in rows:
        sid = r["student_id"]
        # remove every enrollment this student has this year (all terms)
        cur.execute("DELETE FROM enrollments WHERE student_id=? AND academic_year=?", (sid, year))
        # re-enrol in the target class for the whole year (no duplicates)
        for term in ("Term 1", "Term 2", "Term 3"):
            cur.execute("INSERT INTO enrollments(student_id,class_id,term,academic_year) VALUES(?,?,?,?)",
                        (sid, to_id, term, year))
        moved += 1
    cur.commit()
    log_activity("Class promoted", f"{moved} students moved from {from_c['name']} to {to_c['name']} — {from_c['name']} is now empty")
    return jsonify({"moved": moved, "from": from_c["name"], "to": to_c["name"]})

@app.route("/api/students/import", methods=["POST"])
@admin_required
def import_students():
    """Bulk-import students from pasted CSV text (header: first_name,last_name,gender,class,parent_name,parent_phone)."""
    data = request.get_json(force=True) or {}
    text = (data.get("data") or "").strip()
    lines = [l for l in text.splitlines() if l.strip()]
    if len(lines) < 2:
        return jsonify({"error": "Paste the CSV including the header row"}), 400
    s = settings_map()
    term, year = s.get("current_term", "Term 3"), s.get("academic_year", "2026")
    class_cache = {r["name"]: r["id"] for r in q("SELECT id, name FROM classes")}
    created, skipped, errors = 0, 0, []
    for i, line in enumerate(lines[1:], start=2):
        cols = [c.strip().strip('"') for c in line.split(",")]
        if len(cols) < 3:
            errors.append(f"Line {i}: too few columns"); skipped += 1; continue
        fn, ln, gender = cols[0], cols[1], cols[2]
        if not fn or not ln:
            skipped += 1; continue
        gender = gender if gender in ("Male", "Female") else ("Male" if gender.lower().startswith("m") else "Female")
        class_name = cols[3] if len(cols) > 3 else ""
        cid = class_cache.get(class_name)
        if class_name and not cid:
            errors.append(f"Line {i}: class '{class_name}' not found"); skipped += 1; continue
        parent = cols[4] if len(cols) > 4 else ""
        phone = cols[5] if len(cols) > 5 else ""
        adm = f"GF/{year}/{datetime.date.today().year % 100}{random_seq()}"
        try:
            st_id = exe("""INSERT INTO students(admission_no,first_name,last_name,gender,parent_name,parent_phone,status)
                           VALUES(?,?,?,?,?,?,'Active')""",
                        (adm, fn, ln, gender, parent, phone))
        except sqlite3.IntegrityError:
            adm = f"GF/{year}/{datetime.date.today().year % 100}{random_seq()}{random_seq()}"
            st_id = exe("""INSERT INTO students(admission_no,first_name,last_name,gender,parent_name,parent_phone,status)
                           VALUES(?,?,?,?,?,?,'Active')""",
                        (adm, fn, ln, gender, parent, phone))
        if cid:
            exe("INSERT INTO enrollments(student_id,class_id,term,academic_year) VALUES(?,?,?,?)",
                (st_id, cid, term, year))
        created += 1
    log_activity("Bulk import", f"{created} students imported ({skipped} skipped)")
    return jsonify({"created": created, "skipped": skipped, "errors": errors[:20]})

# ------------------------------------------------------------------ teachers
@app.route("/api/teachers")
@login_required
def list_teachers():
    return jsonify(rows_to_dicts(q("""SELECT t.*, s.name subject_name, s.code subject_code
                                      FROM teachers t LEFT JOIN subjects s ON s.id=t.subject_id
                                      WHERE t.active=1 ORDER BY t.first_name""")))

@app.route("/api/teachers", methods=["POST"])
@admin_required
def add_teacher():
    d = request.get_json(force=True) or {}
    if not d.get("first_name") or not d.get("last_name"):
        return jsonify({"error": "Names required"}), 400
    tid = exe("""INSERT INTO teachers(tsc_no,first_name,last_name,gender,phone,email,subject_id,employment_type)
                 VALUES(?,?,?,?,?,?,?,?)""",
              (d.get("tsc_no"), d.get("first_name"), d.get("last_name"), d.get("gender"),
               d.get("phone"), d.get("email"), d.get("subject_id") or None, d.get("employment_type") or "Permanent"))
    return jsonify({"id": tid})

@app.route("/api/teachers/<int:tid>", methods=["PUT"])
@admin_required
def update_teacher(tid):
    d = request.get_json(force=True) or {}
    fields = ["tsc_no", "first_name", "last_name", "gender", "phone", "email", "subject_id", "employment_type"]
    sets = [f for f in fields if f in d]
    if sets:
        exe("UPDATE teachers SET " + ", ".join(f"{f}=?" for f in sets) + " WHERE id=?",
            tuple(d[f] for f in sets) + (tid,))
    return jsonify({"ok": True})

# ------------------------------------------------------------------ classes
@app.route("/api/classes")
@login_required
def list_classes():
    year = settings_map().get("academic_year", "2026")
    rows = q("""SELECT c.*, t.first_name ct_first, t.last_name ct_last,
                (SELECT COUNT(DISTINCT e.student_id) FROM enrollments e WHERE e.class_id=c.id AND e.academic_year=?) cnt
                FROM classes c LEFT JOIN teachers t ON t.id=c.class_teacher_id
                WHERE c.academic_year=? ORDER BY c.grade, c.stream""", (year, year))
    return jsonify(rows_to_dicts(rows))

@app.route("/api/classes", methods=["POST"])
@admin_required
def add_class():
    d = request.get_json(force=True) or {}
    if not d.get("name") or not d.get("grade"):
        return jsonify({"error": "name and grade required"}), 400
    cid = exe("INSERT INTO classes(name,grade,stream,academic_year,capacity,class_teacher_id) VALUES(?,?,?,?,?,?)",
              (d["name"], d["grade"], d.get("stream"), d.get("academic_year") or settings_map().get("academic_year", "2026"),
               d.get("capacity") or 45, d.get("class_teacher_id")))
    log_activity("Class created", d["name"])
    return jsonify({"id": cid})

@app.route("/api/classes/<int:cid>", methods=["PUT"])
@admin_required
def update_class(cid):
    d = request.get_json(force=True) or {}
    fields = ["name", "grade", "stream", "capacity", "class_teacher_id"]
    sets = [f for f in fields if f in d]
    if sets:
        exe("UPDATE classes SET " + ", ".join(f"{f}=?" for f in sets) + " WHERE id=?", tuple(d[f] for f in sets) + (cid,))
    return jsonify({"ok": True})

@app.route("/api/classes/<int:cid>", methods=["DELETE"])
@admin_required
def delete_class(cid):
    c = q1("SELECT * FROM classes WHERE id=?", (cid,))
    if not c:
        return jsonify({"error": "Class not found"}), 404
    cur = db()
    cur.execute("DELETE FROM timetable WHERE class_id=?", (cid,))
    cur.execute("DELETE FROM fee_structures WHERE class_id=?", (cid,))
    cur.execute("DELETE FROM enrollments WHERE class_id=?", (cid,))
    cur.execute("DELETE FROM attendance WHERE class_id=?", (cid,))
    cur.execute("DELETE FROM classes WHERE id=?", (cid,))
    cur.commit()
    log_activity("Class removed", c["name"])
    return jsonify({"ok": True, "removed": c["name"]})

@app.route("/api/classes/<int:cid>/students")
@login_required
def class_students(cid):
    rows = q("""SELECT DISTINCT st.* FROM enrollments e JOIN students st ON st.id=e.student_id
                WHERE e.class_id=? ORDER BY st.first_name""", (cid,))
    return jsonify(rows_to_dicts(rows))

# ------------------------------------------------------------------ subjects
@app.route("/api/subjects")
@login_required
def list_subjects():
    return jsonify(rows_to_dicts(q("""SELECT s.*, t.first_name t_first, t.last_name t_last
                                      FROM subjects s LEFT JOIN teachers t ON t.id=s.teacher_id
                                      ORDER BY s.category, s.name""")))

@app.route("/api/subjects", methods=["POST"])
@admin_required
def add_subject():
    d = request.get_json(force=True) or {}
    if not d.get("name") or not d.get("code"):
        return jsonify({"error": "name and code required"}), 400
    try:
        sid = exe("INSERT INTO subjects(name,code,category,grades,teacher_id) VALUES(?,?,?,?,?)",
                  (d["name"], d["code"].upper(), d.get("category"),
                   d.get("grades") or "1,2,3,4,5,6,7,8,9,10,11,12", d.get("teacher_id")))
    except sqlite3.IntegrityError:
        return jsonify({"error": "Code already exists"}), 400
    return jsonify({"id": sid})

@app.route("/api/subjects/<int:sid>", methods=["PUT"])
@admin_required
def update_subject(sid):
    d = request.get_json(force=True) or {}
    fields = ["name", "code", "category", "grades", "teacher_id"]
    sets = [f for f in fields if f in d]
    if sets:
        exe("UPDATE subjects SET " + ", ".join(f"{f}=?" for f in sets) + " WHERE id=?", tuple(d[f] for f in sets) + (sid,))
    return jsonify({"ok": True})

# ------------------------------------------------------------------ exams & marks
@app.route("/api/exams")
@login_required
def list_exams():
    rows = q("""SELECT e.*, (SELECT COUNT(DISTINCT es.student_id) FROM exam_scores es WHERE es.exam_id=e.id) students
                FROM exams e ORDER BY e.id DESC""")
    return jsonify(rows_to_dicts(rows))

@app.route("/api/exams", methods=["POST"])
@admin_required
def add_exam():
    d = request.get_json(force=True) or {}
    if not d.get("name") or not d.get("term"):
        return jsonify({"error": "name and term required"}), 400
    eid = exe("INSERT INTO exams(name,term,academic_year,status) VALUES(?,?,?,?)",
              (d["name"], d["term"], d.get("academic_year") or settings_map().get("academic_year", "2026"),
               d.get("status") or "Open"))
    return jsonify({"id": eid})

@app.route("/api/exams/<int:eid>", methods=["PUT"])
@admin_required
def update_exam(eid):
    d = request.get_json(force=True) or {}
    fields = ["name", "term", "status"]
    sets = [f for f in fields if f in d]
    if sets:
        exe("UPDATE exams SET " + ", ".join(f"{f}=?" for f in sets) + " WHERE id=?", tuple(d[f] for f in sets) + (eid,))
    return jsonify({"ok": True})

@app.route("/api/exams/<int:eid>")
@login_required
def exam_detail(eid):
    ex = q1("SELECT * FROM exams WHERE id=?", (eid,))
    if not ex:
        return jsonify({"error": "Not found"}), 404
    s = settings_map()
    term = ex["term"]
    year = ex["academic_year"]
    class_rows = q("""SELECT c.id, c.name,
                      (SELECT COUNT(DISTINCT es.student_id) FROM exam_scores es
                       JOIN enrollments e ON e.student_id=es.student_id AND e.term=? AND e.academic_year=?
                       WHERE es.exam_id=? AND e.class_id=c.id) cnt
                      FROM classes c WHERE c.academic_year=? ORDER BY c.grade, c.stream""", (term, year, eid, year))
    return jsonify({"exam": dict(ex),
                    "subjects": rows_to_dicts(q("SELECT * FROM subjects ORDER BY name")),
                    "classes": rows_to_dicts(class_rows)})

@app.route("/api/exams/<int:eid>/marks")
@login_required
def exam_marks(eid):
    class_id = request.args.get("class_id")
    if not class_id:
        return jsonify({"error": "class_id required"}), 400
    students = q("""SELECT st.id, st.admission_no, st.first_name, st.last_name
                    FROM enrollments e JOIN students st ON st.id=e.student_id
                    WHERE e.class_id=? AND e.term=(SELECT term FROM exams WHERE id=?)
                    ORDER BY st.first_name""", (class_id, eid))
    clr = q1("SELECT grade FROM classes WHERE id=?", (class_id,))
    gnum = int(clr["grade"].split()[-1]) if clr else 7
    subjects = subjects_for_grade(gnum)
    scores = q("SELECT student_id, subject_id, score FROM exam_scores WHERE exam_id=?", (eid,))
    score_map = {(r["student_id"], r["subject_id"]): r["score"] for r in scores}
    matrix = []
    for st in students:
        row = {"id": st["id"], "admission_no": st["admission_no"],
               "name": f"{st['first_name']} {st['last_name']}", "scores": {}}
        for su in subjects:
            row["scores"][su["id"]] = score_map.get((st["id"], su["id"]))
        matrix.append(row)
    return jsonify({"students": matrix, "subjects": rows_to_dicts(subjects)})

@app.route("/api/exams/<int:eid>/marks", methods=["POST"])
@academic_required
def save_marks(eid):
    data = request.get_json(force=True) or {}
    updates = data.get("scores") or []
    cur = db()
    n = 0
    for u in updates:
        sid_, subj_, sc = u["student_id"], u["subject_id"], u.get("score")
        if sc is None or sc == "":
            cur.execute("DELETE FROM exam_scores WHERE exam_id=? AND student_id=? AND subject_id=?",
                        (eid, sid_, subj_))
            continue
        sc = float(sc)
        sc = max(0, min(100, sc))
        # all grades use CBC achievement levels
        clr = q1("""SELECT c.grade FROM enrollments e JOIN classes c ON c.id=e.class_id
                    WHERE e.student_id=? AND e.term=(SELECT term FROM exams WHERE id=?)
                    AND e.academic_year=(SELECT academic_year FROM exams WHERE id=?) LIMIT 1""",
                 (sid_, eid, eid))
        scale = scale_for_grade(clr["grade"] if clr else "Grade 7")
        letter, pts = grade_for(sc, scale)
        cur.execute("""INSERT INTO exam_scores(exam_id,student_id,subject_id,score,grade,points)
                       VALUES(?,?,?,?,?,?)
                       ON CONFLICT(exam_id,student_id,subject_id)
                       DO UPDATE SET score=excluded.score, grade=excluded.grade, points=excluded.points""",
                    (eid, sid_, subj_, sc, letter, pts))
        n += 1
    cur.commit()
    return jsonify({"saved": n})

@app.route("/api/exams/<int:eid>/comments")
@academic_required
def exam_comments_get(eid):
    rows = q("SELECT student_id, comment FROM exam_comments WHERE exam_id=?", (eid,))
    return jsonify({r["student_id"]: r["comment"] for r in rows})

@app.route("/api/exams/<int:eid>/comments", methods=["POST"])
@academic_required
def exam_comments_save(eid):
    d = request.get_json(force=True) or {}
    comments = d.get("comments") or {}   # {student_id: text}
    cur = db()
    n = 0
    for sid, text in comments.items():
        text = (text or "").strip()
        if text:
            cur.execute("""INSERT INTO exam_comments(exam_id,student_id,comment) VALUES(?,?,?)
                           ON CONFLICT(exam_id,student_id) DO UPDATE SET comment=excluded.comment""",
                        (eid, sid, text[:400]))
            n += 1
    cur.commit()
    return jsonify({"saved": n})

# ------------------------------------------------------------------ analytics
@app.route("/api/analytics")
@login_required
def analytics():
    exam_id = request.args.get("exam_id", type=int)
    exams = q("SELECT * FROM exams ORDER BY id")
    if not exam_id and exams:
        exam_id = exams[-1]["id"]
    if not exam_id:
        return jsonify({"exams": [], "data": None})
    ex = q1("SELECT * FROM exams WHERE id=?", (exam_id,))
    s = settings_map()
    term = ex["term"]
    year = ex["academic_year"]

    results = exam_results(exam_id)
    student_list = students_with_class(term)
    student_map = {st["id"]: st for st in student_list}

    rows = []
    for sid, res in results.items():
        st = student_map.get(sid)
        if not st:
            continue
        scale = scale_for_grade(st.get("grade") or st.get("class_name") or "Grade 7")
        rows.append({
            "student_id": sid, "admission_no": st["admission_no"],
            "name": f"{st['first_name']} {st['last_name']}",
            "gender": st["gender"], "class_name": st.get("class_name") or "—",
            "subjects": res["subjects"], "total_points": res["total_points"],
            "mean": res["mean"], "avg_pts": res["avg_pts"],
            "scale": scale,
            "mean_grade": mean_grade_from_points(res["avg_pts"], scale),
        })
    for r in rows:
        class_peers = sorted([x for x in rows if x["class_name"] == r["class_name"]],
                             key=lambda x: (-x["avg_pts"], x["name"]))
        r["class_pos"] = next(i + 1 for i, x in enumerate(class_peers) if x["student_id"] == r["student_id"])
        r["class_size"] = len(class_peers)
    ranked = sorted(rows, key=lambda x: (-x["avg_pts"], x["name"]))
    for i, r in enumerate(ranked):
        r["overall_pos"] = i + 1

    grade_dist = {}
    for r in rows:
        grade_dist[r["mean_grade"]] = grade_dist.get(r["mean_grade"], 0) + 1
    # order: CBC levels (E,M,A,B)
    order = ["E", "M", "A", "B", "A-", "B+", "B-", "C+", "C", "C-", "D+", "D", "D-"]
    seen = set()
    ordered = []
    for g in order + ["A", "B"]:
        if g not in seen:
            seen.add(g); ordered.append(g)
    grade_dist = [{"grade": g, "count": grade_dist.get(g, 0)} for g in ordered]

    subject_means = rows_to_dicts(q("""SELECT su.name, su.code, ROUND(AVG(es.score),1) mean,
                                       MAX(es.score) highest, MIN(es.score) lowest,
                                       (SELECT t.first_name FROM teachers t WHERE t.subject_id=su.id) teacher
                                       FROM exam_scores es JOIN subjects su ON su.id=es.subject_id
                                       WHERE es.exam_id=? GROUP BY su.id ORDER BY mean DESC""", (exam_id,)))

    class_means = rows_to_dicts(q("""SELECT c.name, ROUND(AVG(es.score),1) mean, COUNT(DISTINCT es.student_id) students
                                     FROM exam_scores es
                                     JOIN enrollments e ON e.student_id=es.student_id AND e.term=? AND e.academic_year=?
                                     JOIN classes c ON c.id=e.class_id
                                     WHERE es.exam_id=? GROUP BY c.id ORDER BY mean DESC""", (term, year, exam_id)))

    gender_perf = rows_to_dicts(q("""SELECT st.gender, ROUND(AVG(es.score),1) mean, COUNT(DISTINCT es.student_id) students
                                     FROM exam_scores es JOIN students st ON st.id=es.student_id
                                     WHERE es.exam_id=? GROUP BY st.gender""", (exam_id,)))

    trend = rows_to_dicts(q("""SELECT e.id, e.name, e.term, ROUND(AVG(es.score),1) mean, COUNT(DISTINCT es.student_id) students
                               FROM exams e JOIN exam_scores es ON es.exam_id=e.id
                               GROUP BY e.id ORDER BY e.id"""))

    top10 = ranked[:10]
    return jsonify({
        "exams": rows_to_dicts(exams),
        "selected_exam": dict(ex),
        "data": {
            "rows": rows, "ranked": ranked,
            "grade_dist": grade_dist,
            "subject_means": subject_means,
            "class_means": class_means,
            "gender_perf": gender_perf,
            "trend": trend,
            "top10": top10,
            "overall_mean": round(sum(r["mean"] for r in rows) / len(rows), 1) if rows else 0,
        }
    })

@app.route("/api/analytics/student/<int:sid>")
@login_required
def student_analytics(sid):
    exam_id = request.args.get("exam_id", type=int)
    st = q1("SELECT * FROM students WHERE id=?", (sid,))
    if not st:
        return jsonify({"error": "Not found"}), 404
    cl = student_class(sid)
    out = dict(st)
    out["class"] = cl
    exams = q("SELECT * FROM exams ORDER BY id")
    selected = q1("SELECT * FROM exams WHERE id=?", (exam_id,)) if exam_id else exams[-1]
    if not selected:
        return jsonify({"error": "No exams"}), 400

    per_subject = rows_to_dicts(q("""SELECT su.name, es.score, es.grade, es.points,
                                     (SELECT ROUND(AVG(score),1) FROM exam_scores x
                                      WHERE x.exam_id=? AND x.subject_id=su.id) subject_mean
                                     FROM exam_scores es JOIN subjects su ON su.id=es.subject_id
                                     WHERE es.exam_id=? AND es.student_id=?
                                     ORDER BY es.points DESC""", (selected["id"], selected["id"], sid)))
    agg = q1("""SELECT COUNT(*) subjects, SUM(points) total_points, ROUND(AVG(score),1) mean,
                ROUND(AVG(points),2) avg_pts FROM exam_scores WHERE exam_id=? AND student_id=?""",
             (selected["id"], sid))
    term = selected["term"]
    peers = q("""SELECT es.student_id, ROUND(AVG(es.points),2) avg_pts
                 FROM exam_scores es
                 JOIN enrollments e ON e.student_id=es.student_id AND e.term=? AND e.academic_year=?
                 JOIN classes c ON c.id=e.class_id
                 WHERE es.exam_id=? AND c.id=?
                 GROUP BY es.student_id""", (term, selected["academic_year"], selected["id"], cl["id"] if cl else -1))
    peers = sorted(peers, key=lambda r: -r["avg_pts"])
    rank = next((i + 1 for i, p in enumerate(peers) if p["student_id"] == sid), None)

    trend = rows_to_dicts(q("""SELECT e.id, e.term, e.name, ROUND(AVG(es.score),1) mean, ROUND(AVG(es.points),2) avg_pts
                               FROM exams e JOIN exam_scores es ON es.exam_id=e.id AND es.student_id=?
                               GROUP BY e.id ORDER BY e.id""", (sid,)))

    scale = scale_for_grade(cl["grade"] if cl else "Grade 7")
    return jsonify({
        "exams": rows_to_dicts(exams),
        "selected_exam": dict(selected),
        "student": out,
        "scale": scale,
        "per_subject": per_subject,
        "agg": dict(agg) if agg else None,
        "class_rank": rank, "class_size": len(peers),
        "trend": trend,
    })

# ------------------------------------------------------------------ finance
@app.route("/api/payment-types")
@login_required
def payment_types_list():
    rows = q("""SELECT pt.*, (SELECT COUNT(*) FROM fee_payments fp WHERE fp.payment_type_id=pt.id) payments
                FROM payment_types pt ORDER BY pt.active DESC, pt.category, pt.name""")
    return jsonify(rows_to_dicts(rows))

@app.route("/api/payment-types", methods=["POST"])
@admin_required
def add_payment_type():
    d = request.get_json(force=True) or {}
    if not d.get("name"):
        return jsonify({"error": "name required"}), 400
    try:
        pid = exe("INSERT INTO payment_types(name,category,default_amount) VALUES(?,?,?)",
                  (d["name"].strip(), d.get("category") or "Fees", d.get("default_amount")))
    except sqlite3.IntegrityError:
        return jsonify({"error": "A payment type with this name already exists"}), 400
    log_activity("Payment type created", d["name"].strip())
    return jsonify({"id": pid})

@app.route("/api/payment-types/<int:pid>", methods=["PUT"])
@admin_required
def update_payment_type(pid):
    d = request.get_json(force=True) or {}
    fields = ["name", "category", "default_amount", "active"]
    sets = [f for f in fields if f in d]
    if sets:
        exe("UPDATE payment_types SET " + ", ".join(f"{f}=?" for f in sets) + " WHERE id=?",
            tuple(d[f] for f in sets) + (pid,))
    return jsonify({"ok": True})

@app.route("/api/finance")
@login_required
def finance():
    s = settings_map()
    term = request.args.get("term") or s.get("current_term", "Term 3")
    year = request.args.get("year") or s.get("academic_year", "2026")
    billed, paid, class_rows = finance_snapshot(term, year)
    students = []
    for st in students_with_class(term):
        b = billed_for(st["id"], term, year)
        p = paid_for(st["id"], term, year)
        students.append({**st, "billed": b, "paid": p, "balance": b - p,
                         "payments": q1("""SELECT COUNT(*) c FROM fee_payments
                                           WHERE student_id=? AND term=?""", (st["id"], term))["c"]})
    students.sort(key=lambda x: (-x["balance"], x["first_name"]))
    recent = rows_to_dicts(q("""SELECT fp.*, st.first_name, st.last_name, st.admission_no,
                                       pt.name payment_type_name, pt.category payment_type_category
                                FROM fee_payments fp
                                JOIN students st ON st.id=fp.student_id
                                LEFT JOIN payment_types pt ON pt.id=fp.payment_type_id
                                ORDER BY fp.payment_date DESC, fp.id DESC LIMIT 12"""))
    terms = ["Term 1", "Term 2", "Term 3"]
    type_rows = q("""SELECT COALESCE(pt.name,'Other') name, COALESCE(pt.category,'Fees') category,
                            SUM(fp.amount) amount, COUNT(*) count
                     FROM fee_payments fp
                     LEFT JOIN payment_types pt ON pt.id=fp.payment_type_id
                     WHERE fp.term=? AND fp.payment_date LIKE ?
                     GROUP BY pt.id ORDER BY amount DESC""", (term, f"{year}-%"))
    type_breakdown = rows_to_dicts(type_rows)
    return jsonify({"term": term, "year": year, "terms": terms,
                    "billed": billed, "paid": paid, "arrears": billed - paid,
                    "rate": round(paid / billed * 100, 1) if billed else 0,
                    "class_rows": class_rows, "students": students, "recent": recent,
                    "type_breakdown": type_breakdown})

@app.route("/api/finance/payments", methods=["POST"])
@finance_required
def record_payment():
    d = request.get_json(force=True) or {}
    if not d.get("student_id") or not d.get("amount"):
        return jsonify({"error": "student_id and amount required"}), 400
    amount = float(d["amount"])
    if amount <= 0:
        return jsonify({"error": "Amount must be positive"}), 400
    ptid = d.get("payment_type_id")
    if ptid:
        pt = q1("SELECT * FROM payment_types WHERE id=? AND active=1", (ptid,))
        if not pt:
            return jsonify({"error": "Invalid payment type"}), 400
    receipt = "RCP-" + str(q1("SELECT COALESCE(MAX(CAST(SUBSTR(receipt_no,5) AS INTEGER)),9999)+1 m FROM fee_payments")["m"])
    year = settings_map().get("academic_year", "2026")
    term = d.get("term") or settings_map().get("current_term", "Term 3")
    method = d.get("method") or "M-PESA"
    notes = d.get("notes") or ""
    if d.get("auto_split"):
        # one payment, automatically applied to outstanding fees + transport (oldest term first)
        parts = split_allocation(d["student_id"], amount, year)
        if not parts:
            return jsonify({"error": "No outstanding fees to apply this payment to"}), 400
        for part in parts:
            part_notes = f"Auto-split · {part['label']} · {part['term']}" + (f" · {notes}" if notes else "")
            exe("""INSERT INTO fee_payments(student_id,payment_type_id,amount,payment_date,term,method,reference,receipt_no,recorded_by,notes)
                   VALUES(?,?,?,?,?,?,?,?,?,?)""",
                (d["student_id"], part["payment_type_id"], part["amount"],
                 d.get("payment_date") or datetime.date.today().isoformat(),
                 part["term"], method, d.get("reference") or "", receipt, acting_name(), part_notes))
        first_id = q1("SELECT id FROM fee_payments WHERE receipt_no=?", (receipt,))["id"]
        summary = ", ".join(f"{p['label']} {fmt_amount(p['amount'])} ({p['term']})" for p in parts)
        log_activity("Payment recorded (auto-split)", f"{receipt} · {method} · {summary}")
        return jsonify({"id": first_id, "receipt_no": receipt, "split": True, "parts": len(parts)})
    pid = exe("""INSERT INTO fee_payments(student_id,payment_type_id,amount,payment_date,term,method,reference,receipt_no,recorded_by,notes)
                 VALUES(?,?,?,?,?,?,?,?,?,?)""",
              (d["student_id"], ptid, amount, d.get("payment_date") or datetime.date.today().isoformat(),
               term, method, d.get("reference") or "", receipt,
               acting_name(), notes))
    log_activity("Payment recorded", f"{receipt} · {method} · {fmt_amount(amount)}")
    return jsonify({"id": pid, "receipt_no": receipt})

@app.route("/api/finance/payments/<int:pid>", methods=["PUT"])
@finance_required
def update_payment(pid):
    d = request.get_json(force=True) or {}
    rec = q1("SELECT * FROM fee_payments WHERE id=?", (pid,))
    if not rec:
        return jsonify({"error": "Payment not found"}), 404
    rows = q("SELECT * FROM fee_payments WHERE receipt_no=?", (rec["receipt_no"],))
    shared = {}
    for f in ("payment_date", "method", "reference", "notes"):
        if f in d:
            shared[f] = d[f]
    if "amount" in d and d["amount"] is not None:
        new_total = float(d["amount"])
        if new_total <= 0:
            return jsonify({"error": "Amount must be positive"}), 400
        if len(rows) > 1:
            # split receipt: redistribute the new total across the same buckets
            old_total = sum(r["amount"] for r in rows)
            cur = db()
            for r in rows:
                ratio = r["amount"] / old_total if old_total else 1.0 / len(rows)
                cur.execute("UPDATE fee_payments SET amount=? WHERE id=?",
                            (round(new_total * ratio, 2), r["id"]))
            cur.commit()
        else:
            exe("UPDATE fee_payments SET amount=? WHERE id=?", (new_total, pid))
    if shared:
        sets = ", ".join(f"{f}=?" for f in shared)
        cur = db()
        cur.execute("UPDATE fee_payments SET " + sets + " WHERE receipt_no=?", tuple(shared.values()) + (rec["receipt_no"],))
        cur.commit()
    log_activity("Payment updated", f"{rec['receipt_no']}")
    return jsonify({"ok": True})

def statement_data(sid):
    """Itemised billing + payments for a student (shared by staff & portal)."""
    st = q1("SELECT * FROM students WHERE id=?", (sid,))
    if not st:
        return None
    cl = student_class(sid)
    year = settings_map().get("academic_year", "2026")
    billing = []
    for term in ("Term 1", "Term 2", "Term 3"):
        fee = q1("""SELECT amount FROM fee_structures WHERE class_id=? AND term=? AND academic_year=?""",
                 (cl["id"] if cl else -1, term, year))
        route_fee = q1("""SELECT COALESCE(tr.fee,0) f FROM transport_assignments ta
                          JOIN transport_routes tr ON tr.id=ta.route_id
                          WHERE ta.student_id=? AND ta.academic_year=? AND ta.status='Active'""", (sid, year))
        if fee:
            billing.append({"term": term, "academic_year": year, "label": "Tuition", "amount": fee["amount"]})
        if route_fee and route_fee["f"]:
            billing.append({"term": term, "academic_year": year, "label": "Transport", "amount": route_fee["f"]})
    payments = rows_to_dicts(q("""SELECT fp.*, pt.name payment_type_name, pt.category payment_type_category
                                  FROM fee_payments fp
                                  LEFT JOIN payment_types pt ON pt.id=fp.payment_type_id
                                  WHERE fp.student_id=? ORDER BY fp.payment_date DESC, fp.id DESC""", (sid,)))
    total_billed = sum(b["amount"] for b in billing)
    total_paid = sum(p["amount"] for p in payments)
    return {"student": dict(st), "class": cl, "billing": billing, "payments": payments,
            "total_billed": total_billed, "total_paid": total_paid,
            "balance": total_billed - total_paid}

def split_allocation(sid, amount, year):
    """Allocate one payment amount across outstanding term/item buckets.

    Order: oldest term first (Term 1 -> 3); within a term, Transport before
    Tuition. Returns a list of {term, payment_type_id, label, amount}.
    Any remainder beyond the total outstanding becomes a General prepayment
    on the current term so the receipt total always equals the amount paid.
    """
    s = settings_map()
    cl = student_class(sid)
    buckets = []
    ttype = q1("SELECT id FROM payment_types WHERE category='Transport' ORDER BY id LIMIT 1")
    ftype = q1("""SELECT id FROM payment_types WHERE name LIKE '%Tuition%' ORDER BY id LIMIT 1""")
    gtype = q1("SELECT id FROM payment_types WHERE name='General' OR name LIKE '%Prepay%' ORDER BY id LIMIT 1")
    trans_tid = ttype["id"] if ttype else None
    tuit_tid = ftype["id"] if ftype else None
    gen_tid = gtype["id"] if gtype else None

    for term in ("Term 1", "Term 2", "Term 3"):
        fee = q1("SELECT amount FROM fee_structures WHERE class_id=? AND term=? AND academic_year=?",
                 (cl["id"] if cl else -1, term, year))
        route_fee = q1("""SELECT COALESCE(tr.fee,0) f FROM transport_assignments ta
                          JOIN transport_routes tr ON tr.id=ta.route_id
                          WHERE ta.student_id=? AND ta.academic_year=? AND ta.status='Active'""", (sid, year))
        tuition = fee["amount"] if fee else 0
        transport = route_fee["f"] if route_fee else 0
        term_billed = tuition + transport
        if term_billed <= 0:
            continue
        term_paid = q1("""SELECT COALESCE(SUM(amount),0) a FROM fee_payments
                          WHERE student_id=? AND term=? AND payment_date LIKE ?""",
                       (sid, term, f"{year}-%"))["a"]
        term_owed = term_billed - term_paid
        if term_owed <= 0:
            continue
        transport_paid = q1("""SELECT COALESCE(SUM(fp.amount),0) a FROM fee_payments fp
                               WHERE fp.student_id=? AND fp.term=? AND fp.payment_date LIKE ?
                               AND fp.payment_type_id IN (SELECT id FROM payment_types WHERE category='Transport')""",
                            (sid, term, f"{year}-%"))["a"]
        transport_owed = max(0, transport - transport_paid)
        if transport_owed > 0:
            buckets.append({"term": term, "payment_type_id": trans_tid, "label": "Transport",
                            "owed": transport_owed})
        tuition_owed = term_owed - transport_owed
        if tuition_owed > 0:
            buckets.append({"term": term, "payment_type_id": tuit_tid, "label": "Tuition",
                            "owed": tuition_owed})

    remaining = amount
    parts = []
    for b in buckets:
        if remaining <= 0:
            break
        take = min(remaining, b["owed"])
        if take > 0:
            parts.append({"term": b["term"], "payment_type_id": b["payment_type_id"],
                          "label": b["label"], "amount": round(take, 2)})
            remaining -= take
    if remaining > 0:
        parts.append({"term": s.get("current_term", "Term 3"), "payment_type_id": gen_tid,
                      "label": "Prepayment", "amount": round(remaining, 2)})
    return parts

@app.route("/api/finance/statement/<int:sid>")
@login_required
def finance_statement(sid):
    data = statement_data(sid)
    if not data:
        return jsonify({"error": "Not found"}), 404
    return jsonify(data)

@app.route("/api/finance/structure", methods=["POST"])
@finance_required
def set_fee_structure():
    d = request.get_json(force=True) or {}
    if not d.get("class_id") or not d.get("term") or d.get("amount") is None:
        return jsonify({"error": "class_id, term, amount required"}), 400
    year = d.get("academic_year") or settings_map().get("academic_year", "2026")
    exe("""INSERT INTO fee_structures(class_id,term,academic_year,amount) VALUES(?,?,?,?)
           ON CONFLICT(class_id,term,academic_year) DO UPDATE SET amount=excluded.amount""",
        (d["class_id"], d["term"], year, float(d["amount"])))
    return jsonify({"ok": True})

@app.route("/api/finance/reminders", methods=["POST"])
@finance_required
def send_fee_reminders():
    s = settings_map()
    term = s.get("current_term", "Term 3")
    year = s.get("academic_year", "2026")
    sent = 0
    for st in students_with_class(term):
        b = billed_for(st["id"], term, year)
        p = paid_for(st["id"], term, year)
        bal = b - p
        if bal > 0:
            msg = (f"Dear {st.get('parent_name') or 'Parent'}, kindly note that {st['first_name']} {st['last_name']} "
                   f"({st.get('admission_no')}) has a fee balance of {s.get('currency','KSh')} {bal:,.0f} for {term} {year}. "
                   f"Please clear before the exams. - {s.get('school_name')}")
            exe("""INSERT INTO sms_log(to_phone,parent_name,student_name,message,category,status)
                   VALUES(?,?,?,?,?,?)""",
                (st.get("parent_phone"), st.get("parent_name"), f"{st['first_name']} {st['last_name']}", msg,
                 "Fee Reminder", "Queued"))
            sent += 1
    return jsonify({"sent": sent})

# ------------------------------------------------------------------ transport
@app.route("/api/transport/routes")
@login_required
def transport_routes():
    rows = q("""SELECT tr.*,
                (SELECT COUNT(*) FROM transport_assignments ta
                 WHERE ta.route_id=tr.id AND ta.status='Active') assigned
                FROM transport_routes tr ORDER BY tr.name""")
    return jsonify(rows_to_dicts(rows))

@app.route("/api/transport/routes", methods=["POST"])
@admin_required
def add_route():
    d = request.get_json(force=True) or {}
    if not d.get("name"):
        return jsonify({"error": "name required"}), 400
    rid = exe("""INSERT INTO transport_routes(name,route_no,driver_name,driver_phone,capacity,morning_time,evening_time,fee)
                 VALUES(?,?,?,?,?,?,?,?)""",
              (d["name"], d.get("route_no"), d.get("driver_name"), d.get("driver_phone"),
               d.get("capacity") or 40, d.get("morning_time") or "6:30 AM",
               d.get("evening_time") or "4:30 PM", float(d.get("fee") or 0)))
    return jsonify({"id": rid})

@app.route("/api/transport/routes/<int:rid>", methods=["PUT"])
@admin_required
def update_route(rid):
    d = request.get_json(force=True) or {}
    fields = ["name", "route_no", "driver_name", "driver_phone", "capacity", "morning_time", "evening_time", "fee", "status"]
    sets = [f for f in fields if f in d]
    if sets:
        exe("UPDATE transport_routes SET " + ", ".join(f"{f}=?" for f in sets) + " WHERE id=?",
            tuple(d[f] for f in sets) + (rid,))
    return jsonify({"ok": True})

@app.route("/api/transport/assignments")
@login_required
def transport_assignments():
    year = settings_map().get("academic_year", "2026")
    route_id = request.args.get("route_id", type=int)
    assigned = set(r["student_id"] for r in q(
        """SELECT student_id FROM transport_assignments WHERE status='Active' AND academic_year=?""", (year,)))
    rows = []
    for st in students_with_class():
        arows = q("""SELECT ta.route_id r FROM transport_assignments ta
                     WHERE ta.student_id=? AND ta.status='Active' AND ta.academic_year=?""",
                  (st["id"], year))
        assigned_route = arows[0]["r"] if arows else None
        rows.append({"student_id": st["id"], "name": f"{st['first_name']} {st['last_name']}",
                     "admission_no": st["admission_no"], "class_name": st.get("class_name") or "—",
                     "assigned": assigned_route == route_id if route_id else assigned_route is not None})
    return jsonify(rows)

@app.route("/api/transport/assign", methods=["POST"])
@admin_required
def transport_assign():
    d = request.get_json(force=True) or {}
    route_id = d.get("route_id")
    student_ids = d.get("student_ids") or []
    year = settings_map().get("academic_year", "2026")
    cur = db()
    for sid in student_ids:
        cur.execute("""INSERT INTO transport_assignments(student_id,route_id,academic_year,status) VALUES(?,?,?,?)
                       ON CONFLICT(student_id,academic_year) DO UPDATE SET route_id=excluded.route_id, status='Active'""",
                    (sid, route_id, year, "Active"))
    # unassign students no longer in the list for this route
    if student_ids:
        placeholders = ",".join("?" * len(student_ids))
        cur.execute(f"""UPDATE transport_assignments SET status='Inactive'
                        WHERE route_id=? AND academic_year=? AND status='Active'
                        AND student_id NOT IN ({placeholders})""", (route_id, year, *student_ids))
    cur.commit()
    log_activity("Transport updated", f"{len(student_ids)} riders assigned to route {route_id}")
    return jsonify({"assigned": len(student_ids)})

@app.route("/api/transport/register")
@login_required
def transport_register():
    date_ = request.args.get("date") or datetime.date.today().isoformat()
    route_id = request.args.get("route_id", type=int)
    period = request.args.get("period") or "Morning"
    year = settings_map().get("academic_year", "2026")
    if not route_id:
        return jsonify({"error": "route_id required"}), 400
    students = q("""SELECT st.id, st.admission_no, st.first_name, st.last_name
                    FROM transport_assignments ta JOIN students st ON st.id=ta.student_id
                    WHERE ta.route_id=? AND ta.status='Active' AND ta.academic_year=?
                    ORDER BY st.first_name""", (route_id, year))
    existing = {r["student_id"]: r["status"] for r in q(
        "SELECT student_id, status FROM transport_log WHERE date=? AND route_id=? AND period=?",
        (date_, route_id, period))}
    rows = [{**dict(st), "status": existing.get(st["id"], "Boarded")} for st in students]
    return jsonify({"date": date_, "route_id": route_id, "period": period, "rows": rows})

@app.route("/api/transport/register", methods=["POST"])
@academic_required
def save_transport_register():
    d = request.get_json(force=True) or {}
    date_ = d.get("date") or datetime.date.today().isoformat()
    route_id = d.get("route_id")
    period = d.get("period") or "Morning"
    if not route_id:
        return jsonify({"error": "route_id required"}), 400
    cur = db()
    n = 0
    for rec in d.get("records", []):
        cur.execute("""INSERT INTO transport_log(date,route_id,student_id,period,status) VALUES(?,?,?,?,?)
                       ON CONFLICT(date,route_id,student_id,period) DO UPDATE SET status=excluded.status""",
                    (date_, route_id, rec["student_id"], period, rec["status"]))
        n += 1
    cur.commit()
    return jsonify({"saved": n})

@app.route("/api/transport/summary")
@login_required
def transport_summary():
    year = settings_map().get("academic_year", "2026")
    today = datetime.date.today().isoformat()
    routes = q("""SELECT tr.id, tr.name, tr.capacity, tr.fee, tr.driver_name, tr.morning_time, tr.evening_time,
                  (SELECT COUNT(*) FROM transport_assignments ta WHERE ta.route_id=tr.id AND ta.status='Active') assigned,
                  (SELECT COUNT(*) FROM transport_log l WHERE l.route_id=tr.id AND l.date=? AND l.status='Boarded') boarded
                  FROM transport_routes tr WHERE tr.status='Active' ORDER BY tr.name""", (today,))
    return jsonify({
        "routes": rows_to_dicts(routes),
        "total_assigned": q1("SELECT COUNT(*) c FROM transport_assignments WHERE status='Active' AND academic_year=?", (year,))["c"],
        "boarded_today": q1("SELECT COUNT(*) c FROM transport_log WHERE date=? AND status='Boarded'", (today,))["c"],
        "monthly_fees": q1("""SELECT COALESCE(SUM(tr.fee),0) a FROM transport_assignments ta
                              JOIN transport_routes tr ON tr.id=ta.route_id
                              WHERE ta.status='Active' AND ta.academic_year=?""", (year,))["a"],
    })

# ------------------------------------------------------------------ timetable
@app.route("/api/timetable/meta")
@login_required
def timetable_meta():
    return jsonify({"days": DAYS, "periods": PERIODS})

@app.route("/api/timetable")
@login_required
def timetable_grid():
    class_id = request.args.get("class_id", type=int)
    if not class_id:
        return jsonify({"error": "class_id required"}), 400
    entries = q("""SELECT tt.*, s.name subject_name, s.code subject_code, s.category,
                          t.first_name t_first, t.last_name t_last
                   FROM timetable tt
                   LEFT JOIN subjects s ON s.id=tt.subject_id
                   LEFT JOIN teachers t ON t.id=tt.teacher_id
                   WHERE tt.class_id=?
                   ORDER BY tt.day, tt.period""", (class_id,))
    grid = {}
    for e in entries:
        grid.setdefault(e["day"], {})[str(e["period"])] = {
            "subject_id": e["subject_id"], "subject_name": e["subject_name"],
            "subject_code": e["subject_code"], "category": e["category"],
            "teacher_id": e["teacher_id"],
            "teacher_name": ((e["t_first"] or "") + " " + (e["t_last"] or "")).strip(),
        }
    # teacher double-bookings: same teacher in >1 class at same day+period
    conflicts = q("""SELECT day, period, teacher_id FROM timetable
                     WHERE teacher_id IS NOT NULL
                     GROUP BY day, period, teacher_id HAVING COUNT(DISTINCT class_id) > 1""")
    return jsonify({"days": DAYS, "periods": PERIODS, "grid": grid,
                    "conflict_slots": [{"day": r["day"], "period": r["period"]} for r in conflicts]})

@app.route("/api/timetable/teacher")
@login_required
def timetable_teacher():
    u = auth_user()
    tid = request.args.get("teacher_id", type=int)
    if not tid:
        if u["role"] == "teacher" and u.get("teacher_id"):
            tid = u["teacher_id"]
        else:
            return jsonify({"error": "teacher_id required"}), 400
    entries = q("""SELECT tt.*, s.name subject_name, s.code subject_code,
                          c.name class_name
                   FROM timetable tt
                   LEFT JOIN subjects s ON s.id=tt.subject_id
                   LEFT JOIN classes c ON c.id=tt.class_id
                   WHERE tt.teacher_id=?
                   ORDER BY tt.day, tt.period""", (tid,))
    grid = {}
    for e in entries:
        grid.setdefault(e["day"], {})[str(e["period"])] = {
            "class_name": e["class_name"], "subject_name": e["subject_name"],
            "subject_code": e["subject_code"]}
    return jsonify({"days": DAYS, "periods": PERIODS, "grid": grid,
                    "teacher": dict(q1("SELECT * FROM teachers WHERE id=?", (tid,)))})

@app.route("/api/timetable/set", methods=["POST"])
@academic_required
def timetable_set():
    d = request.get_json(force=True) or {}
    class_id = d.get("class_id")
    day = d.get("day")
    period = d.get("period")
    subject_id = d.get("subject_id")
    if not class_id or not day or not period:
        return jsonify({"error": "class_id, day and period required"}), 400
    if day not in DAYS or int(period) not in [p["n"] for p in PERIODS]:
        return jsonify({"error": "Invalid day or period"}), 400
    if not subject_id:
        exe("DELETE FROM timetable WHERE class_id=? AND day=? AND period=?",
            (class_id, day, period))
        return jsonify({"ok": True, "cleared": True})
    subj = q1("SELECT * FROM subjects WHERE id=?", (subject_id,))
    if not subj:
        return jsonify({"error": "Unknown subject"}), 400
    teacher_id = d.get("teacher_id") or subj["teacher_id"]
    if teacher_id:
        other = q1("""SELECT tt.id, c.name FROM timetable tt
                      JOIN classes c ON c.id=tt.class_id
                      WHERE tt.day=? AND tt.period=? AND tt.teacher_id=? AND tt.class_id<>?""",
                   (day, period, teacher_id, class_id))
        if other:
            return jsonify({"error": f"Conflict: this teacher is already booked in {other['name']} "
                                    f"on {day} period {period}"}), 409
    exe("""INSERT INTO timetable(class_id,day,period,subject_id,teacher_id) VALUES(?,?,?,?,?)
           ON CONFLICT(class_id,day,period)
           DO UPDATE SET subject_id=excluded.subject_id, teacher_id=excluded.teacher_id""",
        (class_id, day, period, subject_id, teacher_id))
    return jsonify({"ok": True})

# ------------------------------------------------------------------ receipts
@app.route("/api/finance/receipt/<int:pid>")
@any_required
def receipt_detail(pid):
    p = q1("SELECT * FROM fee_payments WHERE id=?", (pid,))
    if not p:
        return jsonify({"error": "Payment not found"}), 404
    u = auth_user()
    if u["role"] == "guardian" and p["student_id"] not in guardian_student_ids():
        return jsonify({"error": "Forbidden"}), 403
    st = q1("SELECT * FROM students WHERE id=?", (p["student_id"],))
    if not st:
        return jsonify({"error": "Student not found"}), 404
    # aggregate every part sharing this receipt (auto-split payments)
    rows = q("""SELECT fp.*, pt.name payment_type_name, pt.category payment_type_category
                FROM fee_payments fp LEFT JOIN payment_types pt ON pt.id=fp.payment_type_id
                WHERE fp.receipt_no=? ORDER BY fp.id""", (p["receipt_no"],))
    total = sum(r["amount"] for r in rows)
    parts = [{"term": r["term"], "amount": r["amount"],
              "payment_type_name": r["payment_type_name"] or "General",
              "payment_type_category": r["payment_type_category"] or "Other",
              "id": r["id"]} for r in rows]
    cl = student_class(st["id"], rows[0]["term"] if rows else "Term 3")
    s = settings_map()
    paid = q1("SELECT COALESCE(SUM(amount),0) a FROM fee_payments WHERE student_id=?", (st["id"],))["a"]
    billed = 0
    for term in ("Term 1", "Term 2", "Term 3"):
        billed += billed_for(st["id"], term, s.get("academic_year", "2026"))
    pt = q1("SELECT * FROM payment_types WHERE id=?", (p["payment_type_id"],)) if p["payment_type_id"] else None
    main = dict(rows[0])
    main["amount"] = total
    main["split"] = len(rows) > 1
    return jsonify({"payment": main, "student": dict(st), "class": cl,
                    "settings": s, "paid_to_date": paid,
                    "total_billed": billed, "balance": billed - paid,
                    "payment_type": dict(pt) if pt else None,
                    "parts": parts})

# ------------------------------------------------------------------ attendance
@app.route("/api/attendance")
@login_required
def attendance():
    date_ = request.args.get("date") or datetime.date.today().isoformat()
    class_id = request.args.get("class_id")
    s = settings_map()
    term = s.get("current_term", "Term 3")
    rows = []
    if class_id:
        students = q("""SELECT st.id, st.admission_no, st.first_name, st.last_name, st.gender
                        FROM enrollments e JOIN students st ON st.id=e.student_id
                        WHERE e.class_id=? AND e.term=? ORDER BY st.first_name""", (class_id, term))
        existing = q("SELECT student_id, status FROM attendance WHERE date=? AND class_id=?", (date_, class_id))
        emap = {r["student_id"]: r["status"] for r in existing}
        rows = [{**dict(st), "status": emap.get(st["id"], "Present")} for st in students]
    classes = rows_to_dicts(q("SELECT c.id, c.name FROM classes c WHERE c.academic_year=? ORDER BY c.name",
                              (s.get("academic_year", "2026"),)))
    return jsonify({"date": date_, "class_id": class_id, "classes": classes, "rows": rows})

@app.route("/api/attendance", methods=["POST"])
@academic_required
def save_attendance():
    d = request.get_json(force=True) or {}
    date_ = d.get("date") or datetime.date.today().isoformat()
    class_id = d.get("class_id")
    if not class_id:
        return jsonify({"error": "class_id required"}), 400
    cur = db()
    n = 0
    for rec in d.get("records", []):
        cur.execute("""INSERT INTO attendance(date,class_id,student_id,status) VALUES(?,?,?,?)
                       ON CONFLICT(date,class_id,student_id) DO UPDATE SET status=excluded.status""",
                    (date_, class_id, rec["student_id"], rec["status"]))
        n += 1
    cur.commit()
    return jsonify({"saved": n})

@app.route("/api/attendance/summary")
@login_required
def attendance_summary():
    class_id = request.args.get("class_id")
    from_ = request.args.get("from")
    to_ = request.args.get("to") or datetime.date.today().isoformat()
    if not class_id:
        return jsonify({"error": "class_id required"}), 400
    rows = q("""SELECT date, status, COUNT(*) c FROM attendance
                WHERE class_id=? AND (date BETWEEN ? AND ?)
                GROUP BY date, status ORDER BY date""", (class_id, from_ or "2026-01-01", to_))
    days = {}
    for r in rows:
        days.setdefault(r["date"], {"Present": 0, "Absent": 0, "Late": 0, "Permission": 0})
        days[r["date"]][r["status"]] = r["c"]
    return jsonify({"days": [{"date": d, **v} for d, v in sorted(days.items())]})

# ------------------------------------------------------------------ communication
@app.route("/api/announcements")
@any_required
def announcements():
    return jsonify(rows_to_dicts(q("SELECT * FROM announcements ORDER BY id DESC")))

@app.route("/api/announcements", methods=["POST"])
@admin_required
def add_announcement():
    d = request.get_json(force=True) or {}
    if not d.get("title") or not d.get("message"):
        return jsonify({"error": "title and message required"}), 400
    aid = exe("INSERT INTO announcements(title,message,audience,created_by) VALUES(?,?,?,?)",
              (d["title"], d["message"], d.get("audience") or "All", acting_name()))
    log_activity("Announcement posted", d["title"])
    s = settings_map()
    for st in students_with_class():
        if st.get("parent_phone"):
            exe("""INSERT INTO sms_log(to_phone,parent_name,student_name,message,category,status)
                   VALUES(?,?,?,?,?,?)""",
                (st["parent_phone"], st.get("parent_name"), f"{st['first_name']} {st['last_name']}",
                 f"{d['title']}: {d['message'][:140]}", "Announcement", "Queued"))
    return jsonify({"id": aid})

@app.route("/api/smslog")
@login_required
def smslog():
    return jsonify(rows_to_dicts(q("""SELECT * FROM sms_log ORDER BY id DESC LIMIT 100""")))

@app.route("/api/activity")
@login_required
def activity_feed():
    rows = q("""SELECT a.*, u.full_name user_name, u.role user_role
                FROM activity_log a LEFT JOIN users u ON u.id=a.user_id
                ORDER BY a.id DESC LIMIT 12""")
    return jsonify(rows_to_dicts(rows))

# ------------------------------------------------------------------ guardian portal
def guardian_student_ids():
    u = auth_user()
    if not u:
        return []
    rows = q("SELECT student_id FROM guardian_links WHERE user_id=?", (u["id"],))
    return [r["student_id"] for r in rows]

def ensure_guardian(sid):
    return sid in guardian_student_ids()

@app.route("/api/guardian/children")
@guardian_required
def guardian_children():
    ids = guardian_student_ids()
    out = []
    for st in students_with_class():
        if st["id"] in ids:
            out.append({"student_id": st["id"],
                        "name": f"{st['first_name']} {st['last_name']}",
                        "admission_no": st["admission_no"],
                        "class_name": st.get("class_name") or "—",
                        "gender": st["gender"], "profile_pic": st.get("profile_pic")})
    return jsonify(out)

@app.route("/api/guardian/dashboard")
@guardian_required
def guardian_dashboard():
    sid = request.args.get("student_id", type=int)
    if not ensure_guardian(sid):
        return jsonify({"error": "Forbidden"}), 403
    st = q1("SELECT * FROM students WHERE id=?", (sid,))
    s = settings_map()
    term, year = s.get("current_term", "Term 3"), s.get("academic_year", "2026")
    cl = student_class(sid)
    # latest exam with results for this student
    ex = q1("""SELECT e.* FROM exams e WHERE EXISTS (
                 SELECT 1 FROM exam_scores es WHERE es.exam_id=e.id AND es.student_id=?)
               ORDER BY e.id DESC LIMIT 1""", (sid,))
    summary = None
    if ex:
        agg = q1("""SELECT COUNT(*) subjects, SUM(points) total_points, ROUND(AVG(score),1) mean,
                            ROUND(AVG(points),2) avg_pts
                    FROM exam_scores WHERE exam_id=? AND student_id=?""", (ex["id"], sid))
        if agg and agg["subjects"]:
            peers = q("""SELECT es.student_id, ROUND(AVG(es.points),2) avg_pts
                         FROM exam_scores es
                         JOIN enrollments e ON e.student_id=es.student_id AND e.term=? AND e.academic_year=?
                         JOIN classes c ON c.id=e.class_id
                         WHERE es.exam_id=? AND c.id=?
                         GROUP BY es.student_id""", (term, year, ex["id"], cl["id"] if cl else -1))
            peers = sorted(peers, key=lambda r: -r["avg_pts"])
            rank = next((i + 1 for i, p in enumerate(peers) if p["student_id"] == sid), None)
            summary = {"exam_id": ex["id"], "exam_name": ex["name"], "term": ex["term"],
                       "mean": agg["mean"], "total_points": agg["total_points"],
                       "avg_pts": agg["avg_pts"], "subjects": agg["subjects"],
                       "class_rank": rank, "class_size": len(peers)}
    billed = billed_for(sid, term, year)
    paid = paid_for(sid, term, year)
    route = q1("""SELECT tr.* FROM transport_assignments ta
                  JOIN transport_routes tr ON tr.id=ta.route_id
                  WHERE ta.student_id=? AND ta.academic_year=? AND ta.status='Active'""", (sid, year))
    att_days = q("""SELECT date, status FROM attendance
                    WHERE student_id=? ORDER BY date DESC LIMIT 10""", (sid,))
    att_counts = {"Present": 0, "Absent": 0, "Late": 0, "Permission": 0}
    for r in att_days:
        att_counts[r["status"]] = att_counts.get(r["status"], 0) + 1
    ann = rows_to_dicts(q("SELECT * FROM announcements ORDER BY id DESC LIMIT 3"))
    return jsonify({"student": dict(st), "class": cl,
                    "scale": scale_for_grade(cl["grade"] if cl else "Grade 7"),
                    "exam": summary, "billed": billed, "paid": paid, "balance": billed - paid,
                    "transport": dict(route) if route else None,
                    "attendance": {"recent_days": len(att_days), **att_counts},
                    "announcements": ann, "term": term, "year": year})

@app.route("/api/guardian/results")
@guardian_required
def guardian_results():
    sid = request.args.get("student_id", type=int)
    if not ensure_guardian(sid):
        return jsonify({"error": "Forbidden"}), 403
    st = q1("SELECT * FROM students WHERE id=?", (sid,))
    cl = student_class(sid)
    exam_id = request.args.get("exam_id", type=int)
    exams = q("SELECT * FROM exams ORDER BY id")
    selected = q1("SELECT * FROM exams WHERE id=?", (exam_id,)) if exam_id else exams[-1] if exams else None
    if not selected:
        return jsonify({"exams": rows_to_dicts(exams), "selected_exam": None, "per_subject": [], "agg": None}), 200
    per_subject = rows_to_dicts(q("""SELECT su.name, es.score, es.grade, es.points,
                                     (SELECT ROUND(AVG(score),1) FROM exam_scores x
                                      WHERE x.exam_id=? AND x.subject_id=su.id) subject_mean
                                     FROM exam_scores es JOIN subjects su ON su.id=es.subject_id
                                     WHERE es.exam_id=? AND es.student_id=?
                                     ORDER BY es.points DESC""", (selected["id"], selected["id"], sid)))
    agg = q1("""SELECT COUNT(*) subjects, SUM(points) total_points, ROUND(AVG(score),1) mean,
                ROUND(AVG(points),2) avg_pts FROM exam_scores WHERE exam_id=? AND student_id=?""",
             (selected["id"], sid))
    term, year = selected["term"], selected["academic_year"]
    peers = q("""SELECT es.student_id, ROUND(AVG(es.points),2) avg_pts
                 FROM exam_scores es
                 JOIN enrollments e ON e.student_id=es.student_id AND e.term=? AND e.academic_year=?
                 JOIN classes c ON c.id=e.class_id
                 WHERE es.exam_id=? AND c.id=?
                 GROUP BY es.student_id""", (term, year, selected["id"], cl["id"] if cl else -1))
    peers = sorted(peers, key=lambda r: -r["avg_pts"])
    rank = next((i + 1 for i, p in enumerate(peers) if p["student_id"] == sid), None)
    return jsonify({"exams": rows_to_dicts(exams), "selected_exam": dict(selected),
                    "student": dict(st), "scale": scale_for_grade(cl["grade"] if cl else "Grade 7"),
                    "per_subject": per_subject,
                    "agg": dict(agg) if agg and agg["subjects"] else None,
                    "class_rank": rank, "class_size": len(peers)})

@app.route("/api/guardian/statement")
@guardian_required
def guardian_statement():
    sid = request.args.get("student_id", type=int)
    if not ensure_guardian(sid):
        return jsonify({"error": "Forbidden"}), 403
    data = statement_data(sid)
    if not data:
        return jsonify({"error": "Not found"}), 404
    return jsonify(data)

@app.route("/api/guardian/pay", methods=["POST"])
@guardian_required
def guardian_pay():
    d = request.get_json(force=True) or {}
    sid = d.get("student_id")
    if not ensure_guardian(sid):
        return jsonify({"error": "Forbidden"}), 403
    amount = float(d.get("amount") or 0)
    if amount <= 0:
        return jsonify({"error": "Enter a valid amount"}), 400
    ptid = d.get("payment_type_id")
    if ptid:
        pt = q1("SELECT * FROM payment_types WHERE id=? AND active=1", (ptid,))
        if not pt:
            return jsonify({"error": "Invalid payment type"}), 400
    receipt = "RCP-" + str(q1("SELECT COALESCE(MAX(CAST(SUBSTR(receipt_no,5) AS INTEGER)),9999)+1 m FROM fee_payments")["m"])
    import secrets as _s
    ref = d.get("reference") or "MP" + _s.token_hex(4).upper()
    year = settings_map().get("academic_year", "2026")
    if d.get("auto_split"):
        parts = split_allocation(sid, amount, year)
        if not parts:
            return jsonify({"error": "No outstanding fees to apply this payment to"}), 400
        for part in parts:
            exe("""INSERT INTO fee_payments(student_id,payment_type_id,amount,payment_date,term,method,reference,receipt_no,recorded_by,notes)
                   VALUES(?,?,?,?,?,?,?,?,?,?)""",
                (sid, part["payment_type_id"], part["amount"], datetime.date.today().isoformat(),
                 part["term"], "M-PESA", ref, receipt, acting_name(),
                 f"Auto-split · {part['label']} · {part['term']} · Paid via Parent Portal (M-PESA)"))
        first_id = q1("SELECT id FROM fee_payments WHERE receipt_no=?", (receipt,))["id"]
        log_activity("Portal payment (auto-split)", f"{receipt} · {fmt_amount(amount)} · {acting_name()}")
        return jsonify({"id": first_id, "receipt_no": receipt, "split": True, "parts": len(parts)})
    pid = exe("""INSERT INTO fee_payments(student_id,payment_type_id,amount,payment_date,term,method,reference,receipt_no,recorded_by,notes)
                 VALUES(?,?,?,?,?,?,?,?,?,?)""",
              (sid, d.get("payment_type_id"), amount, datetime.date.today().isoformat(),
               settings_map().get("current_term", "Term 3"), "M-PESA", ref, receipt,
               acting_name(), "Paid via Parent Portal (M-PESA)"))
    log_activity("Portal payment", f"{receipt} · {fmt_amount(amount)} · {acting_name()}")
    return jsonify({"id": pid, "receipt_no": receipt})

@app.route("/api/guardian/transport")
@guardian_required
def guardian_transport():
    sid = request.args.get("student_id", type=int)
    if not ensure_guardian(sid):
        return jsonify({"error": "Forbidden"}), 403
    year = settings_map().get("academic_year", "2026")
    route = q1("""SELECT tr.* FROM transport_assignments ta
                  JOIN transport_routes tr ON tr.id=ta.route_id
                  WHERE ta.student_id=? AND ta.academic_year=? AND ta.status='Active'""", (sid, year))
    logs = rows_to_dicts(q("""SELECT date, period, status FROM transport_log
                              WHERE student_id=? ORDER BY date DESC, period DESC LIMIT 12""", (sid,)))
    return jsonify({"route": dict(route) if route else None, "logs": logs})

@app.route("/api/guardian/attendance")
@guardian_required
def guardian_attendance():
    sid = request.args.get("student_id", type=int)
    if not ensure_guardian(sid):
        return jsonify({"error": "Forbidden"}), 403
    cl = student_class(sid)
    rows = q("""SELECT date, status FROM attendance
                WHERE student_id=? AND class_id=?
                ORDER BY date DESC LIMIT 15""", (sid, cl["id"] if cl else -1))
    counts = {"Present": 0, "Absent": 0, "Late": 0, "Permission": 0}
    days = []
    for r in rows:
        counts[r["status"]] = counts.get(r["status"], 0) + 1
        days.append({"date": r["date"], "status": r["status"]})
    days.reverse()
    total = sum(counts.values())
    rate = round((counts["Present"] + counts["Late"] + counts["Permission"]) / total * 100) if total else 0
    return jsonify({"days": days, "counts": counts, "rate": rate, "total": total})

# ------------------------------------------------------------------ curriculum guide
@app.route("/api/curriculum")
@any_required
def curriculum():
    bands = [
        {"key": "lower", "label": "Lower Primary", "grades": "Grade 1 – 3",
         "scale": "CBC Achievement Levels", "gmin": 1, "gmax": 3},
        {"key": "upper", "label": "Upper Primary", "grades": "Grade 4 – 6",
         "scale": "CBC Achievement Levels", "gmin": 4, "gmax": 6},
        {"key": "jss", "label": "Junior Secondary", "grades": "Grade 7 – 9",
         "scale": "CBC Achievement Levels", "gmin": 7, "gmax": 9},
        {"key": "senior", "label": "Senior Secondary", "grades": "Grade 10 – 12",
         "scale": "CBC Achievement Levels", "gmin": 10, "gmax": 12},
    ]
    all_subj = q("SELECT * FROM subjects ORDER BY name")
    for b in bands:
        subj = []
        for s in all_subj:
            try:
                gs = [int(x) for x in (s["grades"] or "").split(",") if x.strip().isdigit()]
            except Exception:
                gs = []
            if any(b["gmin"] <= g <= b["gmax"] for g in gs):
                subj.append({"id": s["id"], "name": s["name"], "code": s["code"], "category": s["category"]})
        b["subjects"] = subj
    return jsonify(bands)

# ------------------------------------------------------------------ discipline / conduct
@app.route("/api/discipline")
@login_required
def discipline_records():
    rec_type = request.args.get("type", "")
    class_id = request.args.get("class_id", type=int)
    q_ = request.args.get("q", "").strip().lower()
    rows = q("""SELECT cr.*, st.first_name, st.last_name, st.admission_no,
                       c.name class_name
                FROM conduct_records cr
                JOIN students st ON st.id=cr.student_id
                LEFT JOIN enrollments e ON e.student_id=st.id AND e.term=(SELECT value FROM settings WHERE key='current_term')
                LEFT JOIN classes c ON c.id=e.class_id
                ORDER BY cr.record_date DESC, cr.id DESC LIMIT 500""")
    out = []
    for r in rows:
        d = dict(r)
        d["student_name"] = f"{d['first_name']} {d['last_name']}"
        if rec_type and d["record_type"] != rec_type:
            continue
        if class_id and d.get("class_name"):
            cl = q1("SELECT id FROM classes WHERE name=?", (d["class_name"],))
            if not cl or cl["id"] != class_id:
                continue
        if q_ and q_ not in (d["student_name"] + " " + d["admission_no"] + " " + (d["category"] or "")).lower():
            continue
        out.append(d)
    return jsonify(out)

@app.route("/api/discipline", methods=["POST"])
@role_required("admin", "teacher")
def discipline_add():
    d = request.get_json(force=True) or {}
    if not d.get("student_id") or d.get("record_type") not in ("Merit", "Demerit"):
        return jsonify({"error": "student_id and record_type (Merit/Demerit) required"}), 400
    if not d.get("category"):
        return jsonify({"error": "Select a category"}), 400
    rid = exe("""INSERT INTO conduct_records(student_id,record_type,category,description,record_date,recorded_by)
                 VALUES(?,?,?,?,?,?)""",
              (d["student_id"], d["record_type"], d["category"], d.get("description"),
               d.get("record_date") or datetime.date.today().isoformat(), acting_name()))
    st = q1("SELECT first_name, last_name FROM students WHERE id=?", (d["student_id"],))
    log_activity("Conduct recorded", f"{d['record_type']} ({d['category']}) for {st['first_name'] if st else ''} {st['last_name'] if st else ''}")
    return jsonify({"id": rid})

@app.route("/api/discipline/<int:rid>", methods=["DELETE"])
@admin_required
def discipline_delete(rid):
    exe("DELETE FROM conduct_records WHERE id=?", (rid,))
    log_activity("Conduct record removed", f"#{rid}")
    return jsonify({"ok": True})

@app.route("/api/discipline/summary")
@login_required
def discipline_summary():
    s = settings_map()
    term_start = s.get("term_start") or "2026-05-01"
    merits = q1("SELECT COUNT(*) c FROM conduct_records WHERE record_type='Merit' AND record_date>=?", (term_start,))["c"]
    demerits = q1("SELECT COUNT(*) c FROM conduct_records WHERE record_type='Demerit' AND record_date>=?", (term_start,))["c"]
    today = q1("SELECT COUNT(*) c FROM conduct_records WHERE record_date=?", (datetime.date.today().isoformat(),))["c"]
    top = q("""SELECT st.id, st.first_name, st.last_name, st.admission_no, c.name class_name,
                      SUM(CASE WHEN cr.record_type='Merit' THEN 1 ELSE 0 END) merits,
                      SUM(CASE WHEN cr.record_type='Demerit' THEN 1 ELSE 0 END) demerits
               FROM conduct_records cr JOIN students st ON st.id=cr.student_id
               LEFT JOIN enrollments e ON e.student_id=st.id AND e.term=(SELECT value FROM settings WHERE key='current_term')
               LEFT JOIN classes c ON c.id=e.class_id
               WHERE cr.record_date>=?
               GROUP BY st.id ORDER BY (merits - demerits) DESC LIMIT 5""", (term_start,))
    return jsonify({"merits": merits, "demerits": demerits, "today": today,
                    "top": rows_to_dicts(top)})

@app.route("/api/discipline/student/<int:sid>")
@login_required
def discipline_student(sid):
    s = settings_map()
    term_start = s.get("term_start") or "2026-05-01"
    term_end = s.get("term_end") or "2026-08-31"
    merits = q1("SELECT COUNT(*) c FROM conduct_records WHERE student_id=? AND record_type='Merit' AND record_date BETWEEN ? AND ?", (sid, term_start, term_end))["c"]
    demerits = q1("SELECT COUNT(*) c FROM conduct_records WHERE student_id=? AND record_type='Demerit' AND record_date BETWEEN ? AND ?", (sid, term_start, term_end))["c"]
    recent = rows_to_dicts(q("""SELECT * FROM conduct_records WHERE student_id=? ORDER BY record_date DESC, id DESC LIMIT 10""", (sid,)))
    return jsonify({"merits": merits, "demerits": demerits, "net": merits - demerits,
                    "rating": conduct_rating(merits - demerits), "recent": recent})

# ------------------------------------------------------------------ events
@app.route("/api/events")
@any_required
def events_list():
    rows = q("SELECT * FROM school_events ORDER BY event_date")
    today = datetime.date.today().isoformat()
    out = []
    for r in rows:
        d = dict(r)
        d["upcoming"] = d["event_date"] >= today
        out.append(d)
    return jsonify(out)

@app.route("/api/events", methods=["POST"])
@admin_required
def events_add():
    d = request.get_json(force=True) or {}
    if not d.get("title") or not d.get("event_date"):
        return jsonify({"error": "title and event_date required"}), 400
    eid = exe("""INSERT INTO school_events(title,description,event_date,category,audience)
                 VALUES(?,?,?,?,?)""",
              (d["title"].strip(), d.get("description"), d["event_date"],
               d.get("category") or "General", d.get("audience") or "All"))
    log_activity("Event created", f"{d['title'].strip()} on {d['event_date']}")
    return jsonify({"id": eid})

@app.route("/api/events/<int:eid>", methods=["PUT"])
@admin_required
def events_update(eid):
    d = request.get_json(force=True) or {}
    fields = ["title", "description", "event_date", "category", "audience"]
    sets = [f for f in fields if f in d]
    if sets:
        exe("UPDATE school_events SET " + ", ".join(f"{f}=?" for f in sets) + " WHERE id=?",
            tuple(d[f] for f in sets) + (eid,))
    log_activity("Event updated", f"#{eid}")
    return jsonify({"ok": True})

@app.route("/api/events/<int:eid>", methods=["DELETE"])
@admin_required
def events_delete(eid):
    exe("DELETE FROM school_events WHERE id=?", (eid,))
    log_activity("Event removed", f"#{eid}")
    return jsonify({"ok": True})

# ------------------------------------------------------------------ ID cards
def card_number(st, year):
    try:
        num = int(st["admission_no"].split("/")[-1])
    except Exception:
        num = st["id"]
    return f"EP-{year}-{num:05d}"

def idcard_payload(sid):
    st = q1("SELECT * FROM students WHERE id=?", (sid,))
    if not st:
        return None
    cl = student_class(sid)
    s = settings_map()
    year = s.get("academic_year", "2026")
    out = dict(st)
    out["class_name"] = cl["name"] if cl else "—"
    out["grade"] = cl["grade"] if cl else "—"
    out["card_no"] = card_number(out, year)
    out["valid_until"] = f"{year}-12-31"
    out["settings"] = s
    return out

@app.route("/api/idcards/student/<int:sid>")
@login_required
def idcard_student(sid):
    data = idcard_payload(sid)
    if not data:
        return jsonify({"error": "Not found"}), 404
    return jsonify(data)

@app.route("/api/idcards/class/<int:cid>")
@login_required
def idcard_class(cid):
    rows = q("""SELECT DISTINCT st.* FROM enrollments e JOIN students st ON st.id=e.student_id
                WHERE e.class_id=? AND st.status='Active' ORDER BY st.first_name""", (cid,))
    cl = q1("SELECT * FROM classes WHERE id=?", (cid,))
    s = settings_map()
    year = s.get("academic_year", "2026")
    out = []
    for r in rows:
        d = dict(r)
        d["class_name"] = cl["name"] if cl else "—"
        d["grade"] = cl["grade"] if cl else "—"
        d["card_no"] = card_number(d, year)
        out.append(d)
    return jsonify({"class": dict(cl) if cl else None, "students": out,
                    "settings": s, "year": year})

# ------------------------------------------------------------------ homework
@app.route("/api/homework")
@any_required
def homework_list():
    class_id = request.args.get("class_id", type=int)
    subject_id = request.args.get("subject_id", type=int)
    status = request.args.get("status", "")
    rows = q("""SELECT hw.*, c.name class_name, s.name subject_name, s.code subject_code
                FROM homework hw
                JOIN classes c ON c.id=hw.class_id
                LEFT JOIN subjects s ON s.id=hw.subject_id
                ORDER BY hw.due_date, hw.id DESC""")
    today = datetime.date.today().isoformat()
    out = []
    for r in rows:
        d = dict(r)
        if class_id and d["class_id"] != class_id:
            continue
        if subject_id and d["subject_id"] != subject_id:
            continue
        overdue = d["due_date"] and d["due_date"] < today
        if status == "Overdue" and not overdue:
            continue
        if status == "Upcoming" and overdue:
            continue
        d["overdue"] = bool(overdue)
        out.append(d)
    return jsonify(out)

@app.route("/api/homework", methods=["POST"])
@role_required("admin", "teacher")
def homework_add():
    d = request.get_json(force=True) or {}
    if not d.get("title") or not d.get("class_id"):
        return jsonify({"error": "title and class_id required"}), 400
    hid = exe("""INSERT INTO homework(class_id,subject_id,title,description,due_date,assigned_by)
                 VALUES(?,?,?,?,?,?)""",
              (d["class_id"], d.get("subject_id"), d["title"].strip(),
               d.get("description"), d.get("due_date"), acting_name()))
    log_activity("Homework assigned", f"{d['title'][:40]} to class #{d['class_id']}")
    return jsonify({"id": hid})

@app.route("/api/homework/<int:hid>", methods=["PUT"])
@role_required("admin", "teacher")
def homework_update(hid):
    d = request.get_json(force=True) or {}
    fields = ["class_id", "subject_id", "title", "description", "due_date"]
    sets = [f for f in fields if f in d]
    if sets:
        exe("UPDATE homework SET " + ", ".join(f"{f}=?" for f in sets) + " WHERE id=?",
            tuple(d[f] for f in sets) + (hid,))
    log_activity("Homework updated", f"#{hid}")
    return jsonify({"ok": True})

@app.route("/api/homework/<int:hid>", methods=["DELETE"])
@role_required("admin", "teacher")
def homework_delete(hid):
    exe("DELETE FROM homework WHERE id=?", (hid,))
    log_activity("Homework removed", f"#{hid}")
    return jsonify({"ok": True})

@app.route("/api/guardian/homework")
@guardian_required
def guardian_homework():
    sid = request.args.get("student_id", type=int)
    if not ensure_guardian(sid):
        return jsonify({"error": "Forbidden"}), 403
    cl = student_class(sid)
    rows = q("""SELECT hw.*, s.name subject_name, s.code subject_code
                FROM homework hw
                LEFT JOIN subjects s ON s.id=hw.subject_id
                WHERE hw.class_id=?
                ORDER BY hw.due_date""", (cl["id"] if cl else -1,))
    today = datetime.date.today().isoformat()
    out = []
    for r in rows:
        d = dict(r)
        d["overdue"] = bool(d["due_date"] and d["due_date"] < today)
        out.append(d)
    return jsonify(out)

# ------------------------------------------------------------------ excel export
@app.route("/api/export/xlsx", methods=["POST"])
@any_required
def export_xlsx():
    """Generic Excel export: {filename, headers:[...], rows:[[...]]} -> .xlsx download."""
    d = request.get_json(force=True) or {}
    filename = (d.get("filename") or "export").replace(".xlsx", "") + ".xlsx"
    headers = d.get("headers") or []
    rows = d.get("rows") or []
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    if headers:
        ws.append(headers)
        hfill = PatternFill("solid", fgColor="14532D")
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = hfill
            cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(["" if v is None else v for v in row])
    # auto width
    for c in range(1, len(headers) + 1):
        mx = len(str(headers[c - 1])) if headers else 10
        for r in range(2, min(len(rows) + 2, 60)):
            v = str(ws.cell(row=r, column=c).value or "")
            mx = max(mx, len(v))
        ws.column_dimensions[get_column_letter(c)].width = min(mx + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import send_file
    resp = send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)
    return resp

# ------------------------------------------------------------------ payroll
def payroll_settings():
    s = settings_map()
    return {
        "paye_low": float(s.get("paye_low", 24000)), "paye_mid": float(s.get("paye_mid", 32333)),
        "paye_low_rate": float(s.get("paye_low_rate", 0.10)), "paye_mid_rate": float(s.get("paye_mid_rate", 0.25)),
        "paye_high_rate": float(s.get("paye_high_rate", 0.30)), "personal_relief": float(s.get("personal_relief", 2400)),
        "shif_rate": float(s.get("shif_rate", 0.0275)), "shif_enabled": s.get("shif_enabled", "1") != "0",
        "nssf_rate": float(s.get("nssf_rate", 0.06)), "nssf_cap": float(s.get("nssf_cap", 36000)),
        "nssf_enabled": s.get("nssf_enabled", "1") != "0",
        "housing_rate": float(s.get("housing_rate", 0.015)), "housing_enabled": s.get("housing_enabled", "1") != "0",
    }

def compute_payslip(basic, allowances, other_ded=0, ps=None):
    ps = ps or payroll_settings()
    gross = (basic or 0) + (allowances or 0)
    taxable = gross
    paye = 0.0
    paye += min(taxable, ps["paye_low"]) * ps["paye_low_rate"]
    if taxable > ps["paye_low"]:
        paye += (min(taxable, ps["paye_mid"]) - ps["paye_low"]) * ps["paye_mid_rate"]
    if taxable > ps["paye_mid"]:
        paye += (taxable - ps["paye_mid"]) * ps["paye_high_rate"]
    paye = max(0, paye - ps["personal_relief"])
    shif = round(gross * ps["shif_rate"], 2) if ps["shif_enabled"] else 0
    nssf = round(min(gross, ps["nssf_cap"]) * ps["nssf_rate"], 2) if ps["nssf_enabled"] else 0
    housing = round(gross * ps["housing_rate"], 2) if ps["housing_enabled"] else 0
    total_d = round(paye + shif + nssf + housing + (other_ded or 0), 2)
    return {"basic": round(basic or 0, 2), "allowances": round(allowances or 0, 2), "gross": round(gross, 2),
            "paye": round(paye, 2), "shif": shif, "nssf": nssf, "housing": housing,
            "other_deductions": round(other_ded or 0, 2), "total_deductions": total_d,
            "net_pay": round(gross - total_d, 2)}

@app.route("/api/payroll/employees")
@finance_required
def payroll_employees():
    rows = q("""SELECT t.*, s.name subject_name FROM teachers t
                LEFT JOIN subjects s ON s.id=t.subject_id
                WHERE t.active=1 ORDER BY t.first_name""")
    return jsonify(rows_to_dicts(rows))

@app.route("/api/payroll/employees/<int:tid>", methods=["PUT"])
@finance_required
def payroll_employee_update(tid):
    d = request.get_json(force=True) or {}
    fields = ["basic_salary", "allowances", "kra_pin", "nssf_no", "shif_no", "bank_name", "bank_account"]
    sets = [f for f in fields if f in d]
    if sets:
        exe("UPDATE teachers SET " + ", ".join(f"{f}=?" for f in sets) + " WHERE id=?",
            tuple(d[f] for f in sets) + (tid,))
    log_activity("Payroll profile updated", f"teacher #{tid}")
    return jsonify({"ok": True})

@app.route("/api/payroll/settings", methods=["GET", "PUT"])
@finance_required
def payroll_settings_ep():
    if request.method == "GET":
        return jsonify(payroll_settings())
    d = request.get_json(force=True) or {}
    cur = db()
    for k, v in d.items():
        cur.execute("INSERT INTO settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                    (k, str(v)))
    cur.commit()
    log_activity("Payroll settings updated", "")
    return jsonify({"ok": True})

@app.route("/api/payroll/run", methods=["POST"])
@finance_required
def payroll_run():
    d = request.get_json(force=True) or {}
    month = int(d.get("month") or datetime.date.today().month)
    year = int(d.get("year") or datetime.date.today().year)
    months = ["", "January", "February", "March", "April", "May", "June", "July",
              "August", "September", "October", "November", "December"]
    label = f"{months[month]} {year}"
    existing = q1("SELECT id FROM payroll_runs WHERE month=? AND year=?", (month, year))
    if existing:
        return jsonify({"error": f"Payroll for {label} already exists"}), 400
    ps = payroll_settings()
    cur = db()
    rid = cur.execute("""INSERT INTO payroll_runs(month,year,period_label,status,prepared_by)
                         VALUES(?,?,?,'Draft',?)""", (month, year, label, acting_name())).lastrowid
    count = 0
    for t in cur.execute("SELECT id, basic_salary, allowances FROM teachers WHERE active=1"):
        c = compute_payslip(t["basic_salary"], t["allowances"], 0, ps)
        cur.execute("""INSERT INTO payroll_payslips(run_id,teacher_id,basic_salary,allowances,gross,paye,shif,nssf,housing,
                                                    other_deductions,total_deductions,net_pay,paid)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?,0)""",
                    (rid, t["id"], c["basic"], c["allowances"], c["gross"], c["paye"], c["shif"], c["nssf"],
                     c["housing"], c["other_deductions"], c["total_deductions"], c["net_pay"]))
        count += 1
    cur.commit()
    log_activity("Payroll run created", f"{label} · {count} employees · Draft")
    return jsonify({"run_id": rid, "count": count, "label": label})

@app.route("/api/payroll/runs")
@finance_required
def payroll_runs():
    rows = q("""SELECT r.*,
                       (SELECT COUNT(*) FROM payroll_payslips p WHERE p.run_id=r.id) employees,
                       (SELECT SUM(net_pay) FROM payroll_payslips p WHERE p.run_id=r.id) net_total,
                       (SELECT SUM(gross) FROM payroll_payslips p WHERE p.run_id=r.id) gross_total,
                       (SELECT SUM(total_deductions) FROM payroll_payslips p WHERE p.run_id=r.id) deductions_total
                FROM payroll_runs r ORDER BY r.year DESC, r.month DESC""")
    return jsonify(rows_to_dicts(rows))

@app.route("/api/payroll/run/<int:rid>")
@finance_required
def payroll_run_detail(rid):
    run = q1("SELECT * FROM payroll_runs WHERE id=?", (rid,))
    if not run:
        return jsonify({"error": "Not found"}), 404
    slips = q("""SELECT p.*, t.first_name, t.last_name, t.tsc_no, t.bank_name, t.bank_account
                 FROM payroll_payslips p JOIN teachers t ON t.id=p.teacher_id
                 WHERE p.run_id=? ORDER BY t.first_name""", (rid,))
    return jsonify({"run": dict(run), "slips": rows_to_dicts(slips)})

@app.route("/api/payroll/run/<int:rid>/pay", methods=["POST"])
@finance_required
def payroll_run_pay(rid):
    cur = db()
    cur.execute("UPDATE payroll_payslips SET paid=1 WHERE run_id=?", (rid,))
    cur.execute("UPDATE payroll_runs SET status='Paid' WHERE id=?", (rid,))
    cur.commit()
    log_activity("Payroll marked paid", f"run #{rid}")
    return jsonify({"ok": True})

@app.route("/api/payroll/run/<int:rid>", methods=["DELETE"])
@finance_required
def payroll_run_delete(rid):
    exe("DELETE FROM payroll_payslips WHERE run_id=?", (rid,))
    exe("DELETE FROM payroll_runs WHERE id=?", (rid,))
    log_activity("Payroll run deleted", f"run #{rid}")
    return jsonify({"ok": True})

@app.route("/api/payroll/payslip/<int:pid>")
@finance_required
def payroll_payslip(pid):
    p = q1("""SELECT p.*, t.first_name, t.last_name, t.tsc_no, t.kra_pin, t.nssf_no, t.shif_no,
                     t.bank_name, t.bank_account, t.phone, r.period_label, r.month, r.year
              FROM payroll_payslips p
              JOIN teachers t ON t.id=p.teacher_id
              JOIN payroll_runs r ON r.id=p.run_id
              WHERE p.id=?""", (pid,))
    if not p:
        return jsonify({"error": "Not found"}), 404
    out = dict(p)
    out["school"] = settings_map()
    out["amount_words"] = amount_in_words(p["net_pay"])
    return jsonify(out)

@app.route("/api/payroll/summary")
@finance_required
def payroll_summary():
    s = settings_map()
    employees = q1("SELECT COUNT(*) c FROM teachers WHERE active=1")["c"]
    gross_total = q1("SELECT COALESCE(SUM(basic_salary+allowances),0) a FROM teachers WHERE active=1")["a"]
    last_run = q1("""SELECT * FROM payroll_runs ORDER BY year DESC, month DESC LIMIT 1""")
    last_paid = 0
    if last_run:
        last_paid = q1("SELECT COALESCE(SUM(net_pay),0) a FROM payroll_payslips WHERE run_id=?", (last_run["id"],))["a"]
    return jsonify({"employees": employees, "monthly_gross": gross_total,
                    "last_run": dict(last_run) if last_run else None,
                    "last_net": last_paid,
                    "settings": payroll_settings()})

# ------------------------------------------------------------------ library
@app.route("/api/library/books")
@any_required
def library_books():
    q_ = request.args.get("q", "").strip().lower()
    cat = request.args.get("category", "")
    rows = q("""SELECT b.*,
                       (SELECT COUNT(*) FROM book_issues i WHERE i.book_id=b.id AND i.status IN ('Issued','Overdue')) out_count
                FROM books b ORDER BY b.title""")
    out = []
    for r in rows:
        b = dict(r)
        if q_ and q_ not in (b["title"] + " " + (b["author"] or "") + " " + (b["isbn"] or "")).lower():
            continue
        if cat and b["category"] != cat:
            continue
        b["status"] = "Out" if b["available_copies"] <= 0 else ("Low" if b["available_copies"] <= 2 else "Available")
        out.append(b)
    return jsonify(out)

@app.route("/api/library/books", methods=["POST"])
@library_required
def library_add_book():
    d = request.get_json(force=True) or {}
    if not d.get("title"):
        return jsonify({"error": "Book title required"}), 400
    copies = max(1, int(d.get("total_copies") or 1))
    bid = exe("""INSERT INTO books(title,author,isbn,publisher,category,year,total_copies,available_copies,shelf)
                 VALUES(?,?,?,?,?,?,?,?,?)""",
              (d["title"].strip(), d.get("author"), d.get("isbn"), d.get("publisher"),
               d.get("category") or "Textbook", d.get("year"), copies, copies, d.get("shelf")))
    log_activity("Book added", f"{d['title'].strip()} ({copies} copies)")
    return jsonify({"id": bid})

@app.route("/api/library/books/<int:bid>", methods=["PUT"])
@library_required
def library_update_book(bid):
    d = request.get_json(force=True) or {}
    fields = ["title", "author", "isbn", "publisher", "category", "year", "shelf"]
    sets = [f for f in fields if f in d]
    if "total_copies" in d:
        cur = db()
        out = q1("SELECT COUNT(*) c FROM book_issues WHERE book_id=? AND status IN ('Issued','Overdue')", (bid,))["c"]
        new_total = max(int(d["total_copies"]), out)
        cur.execute("UPDATE books SET total_copies=?, available_copies=? WHERE id=?", (new_total, new_total - out, bid))
        cur.commit()
    if sets:
        exe("UPDATE books SET " + ", ".join(f"{f}=?" for f in sets) + " WHERE id=?", tuple(d[f] for f in sets) + (bid,))
    log_activity("Book updated", f"#{bid}")
    return jsonify({"ok": True})

@app.route("/api/library/books/<int:bid>", methods=["DELETE"])
@library_required
def library_delete_book(bid):
    out = q1("SELECT COUNT(*) c FROM book_issues WHERE book_id=? AND status IN ('Issued','Overdue')", (bid,))["c"]
    if out:
        return jsonify({"error": "Cannot delete — copies are currently issued out"}), 400
    exe("DELETE FROM book_issues WHERE book_id=?", (bid,))
    b = q1("SELECT * FROM books WHERE id=?", (bid,))
    exe("DELETE FROM books WHERE id=?", (bid,))
    log_activity("Book removed", b["title"] if b else f"#{bid}")
    return jsonify({"ok": True})

@app.route("/api/library/issues")
@any_required
def library_issues():
    status = request.args.get("status", "")
    rows = q("""SELECT i.*, b.title book_title, b.author book_author, b.category book_category,
                       st.first_name, st.last_name, st.admission_no, c.name class_name
                FROM book_issues i
                JOIN books b ON b.id=i.book_id
                JOIN students st ON st.id=i.student_id
                LEFT JOIN enrollments e ON e.student_id=st.id AND e.term=(SELECT value FROM settings WHERE key='current_term')
                LEFT JOIN classes c ON c.id=e.class_id
                ORDER BY i.id DESC LIMIT 200""")
    out = []
    for r in rows:
        d = dict(r)
        d["student_name"] = f"{d['first_name']} {d['last_name']}"
        if status and d["status"] != status:
            continue
        out.append(d)
    return jsonify(out)

@app.route("/api/library/issue", methods=["POST"])
@library_required
def library_issue():
    d = request.get_json(force=True) or {}
    bid, sid = d.get("book_id"), d.get("student_id")
    if not bid or not sid:
        return jsonify({"error": "book_id and student_id required"}), 400
    book = q1("SELECT * FROM books WHERE id=?", (bid,))
    st = q1("SELECT * FROM students WHERE id=?", (sid,))
    if not book or not st:
        return jsonify({"error": "Book or student not found"}), 404
    if book["available_copies"] <= 0:
        return jsonify({"error": "No copies of this book are currently available"}), 400
    due = d.get("due_date") or (datetime.date.today() + datetime.timedelta(days=14)).isoformat()
    iid = exe("""INSERT INTO book_issues(book_id,student_id,issue_date,due_date,status,issued_by,notes)
                 VALUES(?,?,?,?,?,?,?)""",
              (bid, sid, datetime.date.today().isoformat(), due, "Issued", acting_name(), d.get("notes")))
    exe("UPDATE books SET available_copies=available_copies-1 WHERE id=?", (bid,))
    log_activity("Book issued", f"{book['title']} -> {st['first_name']} {st['last_name']} (due {due})")
    return jsonify({"id": iid})

@app.route("/api/library/return/<int:iid>", methods=["POST"])
@library_required
def library_return(iid):
    rec = q1("SELECT * FROM book_issues WHERE id=?", (iid,))
    if not rec:
        return jsonify({"error": "Issue record not found"}), 404
    if rec["status"] in ("Returned",):
        return jsonify({"error": "This book was already returned"}), 400
    exe("UPDATE book_issues SET return_date=?, status='Returned' WHERE id=?",
        (datetime.date.today().isoformat(), iid))
    exe("UPDATE books SET available_copies=available_copies+1 WHERE id=?", (rec["book_id"],))
    log_activity("Book returned", f"issue #{iid}")
    return jsonify({"ok": True})

@app.route("/api/library/summary")
@any_required
def library_summary():
    total = q1("SELECT COUNT(*) c FROM books")["c"]
    copies = q1("SELECT COALESCE(SUM(total_copies),0) c FROM books")["c"]
    available = q1("SELECT COALESCE(SUM(available_copies),0) c FROM books")["c"]
    issued = q1("SELECT COUNT(*) c FROM book_issues WHERE status IN ('Issued','Overdue')")["c"]
    overdue = q1("SELECT COUNT(*) c FROM book_issues WHERE status='Overdue'")["c"]
    return jsonify({"total_titles": total, "total_copies": copies, "available": available,
                    "issued": issued, "overdue": overdue})

# ------------------------------------------------------------------ multi-school management (super admin)
@app.route("/api/schools")
@super_required
def schools_list():
    out = []
    for sch in list_schools():
        try:
            c = sqlite3.connect(sch["db_path"])
            c.row_factory = sqlite3.Row
            students = c.execute("SELECT COUNT(*) c FROM students WHERE status='Active'").fetchone()["c"]
            teachers = c.execute("SELECT COUNT(*) c FROM teachers WHERE active=1").fetchone()["c"]
            c.close()
        except Exception:
            students = teachers = 0
        out.append({**sch, "students": students, "teachers": teachers})
    return jsonify(out)

@app.route("/api/schools", methods=["POST"])
@super_required
def schools_create():
    d = request.get_json(force=True) or {}
    name = (d.get("name") or "").strip()
    slug = (d.get("slug") or "").strip().lower()
    import re as _re
    if not name:
        return jsonify({"error": "School name required"}), 400
    if not _re.match(r"^[a-z0-9][a-z0-9-]{2,30}$", slug):
        return jsonify({"error": "Slug must be 3-31 chars: lowercase letters, numbers, hyphens"}), 400
    if school_by_slug(slug):
        return jsonify({"error": "A school with this slug already exists"}), 400
    sample = bool(d.get("sample"))
    path = os.path.join(DATA_DIR, f"school_{slug}.db")
    admin_user = (d.get("admin_user") or "admin").strip()
    admin_pass = d.get("admin_pass") or "admin123"
    try:
        import seed as _seed
        _seed.seed_db(path, school_name=name, sample=sample,
                      admin_user=admin_user, admin_pass=admin_pass)
    except Exception as e:
        return jsonify({"error": f"Could not create school: {e}"}), 500
    import seed as _seed2
    _seed2.register_school(slug, name, os.path.relpath(path, BASE_DIR))
    log_activity("School created", f"{name} ({slug})")
    return jsonify({"ok": True, "slug": slug, "name": name, "sample": sample})

@app.route("/api/schools/<int:sid>", methods=["PUT"])
@super_required
def schools_update(sid):
    d = request.get_json(force=True) or {}
    m = open_meta()
    if "active" in d:
        m.execute("UPDATE schools SET active=? WHERE id=?", (1 if d["active"] else 0, sid))
    if "name" in d and d["name"]:
        m.execute("UPDATE schools SET name=? WHERE id=?", (d["name"].strip(), sid))
    m.commit(); m.close()
    return jsonify({"ok": True})

# ------------------------------------------------------------------ settings
@app.route("/api/settings", methods=["GET", "PUT"])
@any_required
def settings_endpoint():
    if request.method == "GET":
        return jsonify(settings_map())
    if auth_user()["role"] != "admin":
        return jsonify({"error": "Forbidden"}), 403
    d = request.get_json(force=True) or {}
    cur = db()
    for k, v in d.items():
        cur.execute("INSERT INTO settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                    (k, str(v)))
    cur.commit()
    return jsonify({"ok": True})

if __name__ == "__main__":
    boot()
    port = int(os.environ.get("PORT", 8000))
    app.run(host="0.0.0.0", port=port, debug=False)
